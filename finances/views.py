from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, F, ExpressionWrapper, DecimalField, Avg
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, logout, authenticate
from datetime import timedelta, datetime, date
import calendar
import csv
import json
import logging
from .models import Transaction, Category, Budget, Account, DebitAccount, CreditAccount, Wallet, SubCategory, Debt, ScheduledTransaction, DashboardPreference
from .forms import ExportForm, ImportForm, TransactionForm, CategoryForm, BudgetForm, DateRangeForm, DebitAccountForm, CreditAccountForm, WalletForm, SubCategoryForm, DebtForm, ScheduledTransactionForm
from django import forms
from django.contrib.auth.models import User
from django.db import models
from decimal import Decimal
import openpyxl
import decimal
from .utils import get_transactions_with_scheduled, generate_scheduled_transactions

def home(request):
    return render(request, 'index.html')

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Welcome back, {username}!')
                return redirect('dashboard')
            else:
                messages.error(request, 'Invalid username or password.')
                return render(request, 'index.html', {'open_login_modal': True, 'form_errors': 'Invalid username or password.', 'form_type': 'login'})
        else:
            # Pass form errors to the modal, flattening them
            from itertools import chain
            errors = list(chain.from_iterable(form.errors.values()))
            return render(request, 'index.html', {'open_login_modal': True, 'form_errors': '<br>'.join(errors), 'form_type': 'login'})
    else:
        form = AuthenticationForm()
        return render(request, 'index.html', {'open_login_modal': True, 'form_type': 'login'})

class CustomUserCreationForm(UserCreationForm):
    email = forms.EmailField(required=True)
    
    class Meta:
        model = User
        fields = ("username", "email", "password1", "password2")
    
    def save(self, commit=True):
        user = super().save(commit=False)
        user.email = self.cleaned_data["email"]
        if commit:
            user.save()
        return user

def signup(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            
            # Create default income categories
            income_categories = [
                {'name': 'Salary/Wages', 'icon': 'bi-cash-coin', 'type': 'income'},
                {'name': 'Business Income', 'icon': 'bi-shop', 'type': 'income'},
                {'name': 'Investments', 'icon': 'bi-graph-up-arrow', 'type': 'income'},
                {'name': 'Other Income', 'icon': 'bi-wallet2', 'type': 'income'},
            ]
            
            # Create default expense categories
            expense_categories = [
                {'name': 'Housing', 'icon': 'bi-house', 'type': 'expense'},
                {'name': 'Utilities', 'icon': 'bi-lightning', 'type': 'expense'},
                {'name': 'Food', 'icon': 'bi-cup-hot', 'type': 'expense'},
                {'name': 'Transportation', 'icon': 'bi-truck', 'type': 'expense'},
                {'name': 'Insurance', 'icon': 'bi-shield', 'type': 'expense'},
                {'name': 'Entertainment', 'icon': 'bi-music-note-beamed', 'type': 'expense'},
                {'name': 'Healthcare', 'icon': 'bi-heart-pulse', 'type': 'expense'},
                {'name': 'Debt Payments', 'icon': 'bi-credit-card', 'type': 'expense'},
                {'name': 'Savings/Investments', 'icon': 'bi-piggy-bank', 'type': 'expense'},
                {'name': 'Miscellaneous', 'icon': 'bi-three-dots', 'type': 'expense'},
            ]
            
            # Create all categories
            for category_data in income_categories + expense_categories:
                Category.objects.create(
                    name=category_data['name'],
                    icon=category_data['icon'],
                    type=category_data['type'],
                    user=user
                )
            
            login(request, user)
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            messages.success(request, f'Account created successfully! Welcome to Budget Tracker.')
            return redirect('dashboard')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                from itertools import chain
                errors = list(chain.from_iterable(form.errors.values()))
                return JsonResponse({'success': False, 'errors': '<br>'.join(errors)})
            # If not AJAX, render home with modal open and errors
            from itertools import chain
            errors = list(chain.from_iterable(form.errors.values()))
            return render(request, 'index.html', {'open_signup_modal': True, 'form_errors': '<br>'.join(errors), 'form_type': 'register'})
    else:
        form = CustomUserCreationForm()
        # If not AJAX, render home with modal open and empty form
        return render(request, 'index.html', {'open_signup_modal': True, 'form_type': 'register'})

@login_required
def scheduled_transactions(request):
    # Get the selected month from the request, default to current month
    selected_month_str = request.GET.get('month', timezone.now().strftime('%Y-%m'))
    selected_month = timezone.make_aware(datetime.strptime(selected_month_str, '%Y-%m'))
    
    # Calculate first and last day of the selected month
    first_day = selected_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if selected_month.month == 12:
        last_day = selected_month.replace(year=selected_month.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        last_day = selected_month.replace(month=selected_month.month + 1, day=1) - timedelta(days=1)
    last_day = last_day.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Get search query
    search_query = request.GET.get('search', '')

    # Get only scheduled transactions for the month (all statuses)
    monthly_transactions = generate_scheduled_transactions(request.user, first_day, last_day)
    # Sort by date, then by status (scheduled first, then failed, then completed)
    def status_order(status):
        if status == 'scheduled':
            return 0
        elif status == 'failed':
            return 1
        elif status == 'completed':
            return 2
        return 3
    monthly_transactions.sort(key=lambda t: (t['date'], status_order(t['status'])))
    
    # Apply search filter if provided
    if search_query:
        monthly_transactions = [
            t for t in monthly_transactions 
            if search_query.lower() in t['name'].lower() or 
               (t.get('note') and search_query.lower() in t['note'].lower())
        ]
    
    # Calculate totals
    expected_income = Decimal('0')
    expected_expense = Decimal('0')
    
    for transaction in monthly_transactions:
        if transaction['status'] == 'failed':
            continue  # Skip failed entries in summary
        if transaction['type'] == 'income':
            expected_income += Decimal(str(transaction['amount']))
        else:
            expected_expense += Decimal(str(transaction['amount']))
    
    # Calculate net sum
    net_sum = expected_income - expected_expense
    
    # Get categories and accounts for the modal form
    categories = Category.objects.filter(user=request.user)
    accounts = Account.objects.filter(user=request.user)

    context = {
        'selected_month': selected_month,
        'search_query': search_query,
        'monthly_transactions': monthly_transactions,
        'income_sum': expected_income,
        'expense_sum': expected_expense,
        'net_sum': net_sum,
        'categories': categories,
        'accounts': accounts
    }
    
    return render(request, 'finances/scheduled_transactions.html', context)

@login_required
def create_scheduled_transaction(request):
    if request.method == 'POST':
        # Check if the request is missing the 'repeats' field but has 'repeat_type'
        post_data = request.POST.copy()
        if 'repeat_type' in post_data and post_data.get('repeat_type') == 'once' and 'repeats' not in post_data:
            post_data['repeats'] = '1'
            
        form = ScheduledTransactionForm(post_data, user=request.user)
        if form.is_valid():
            scheduled_transaction = form.save(commit=False)
            scheduled_transaction.user = request.user
            scheduled_transaction.is_recurring = scheduled_transaction.repeats == 0
            try:
                scheduled_transaction.save()
            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'errors': {'': [str(e)]}
                    }, status=400)
                messages.error(request, f'Error saving scheduled transaction: {e}')
                return render(request, 'finances/scheduled_transaction_form.html', {
                    'form': form,
                    'title': 'Create Scheduled Transaction'
                })
            # Check if it's an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Scheduled transaction created successfully.',
                    'id': scheduled_transaction.id
                })
            messages.success(request, 'Scheduled transaction created successfully.')
            return redirect('scheduled_transactions')
        else:
            # For AJAX requests, return errors as JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = {}
                for field, error_list in form.errors.items():
                    errors[field] = [str(error) for error in error_list]
                # Also include non-field errors
                if form.non_field_errors():
                    errors['__all__'] = [str(e) for e in form.non_field_errors()]
                return JsonResponse({
                    'success': False,
                    'errors': errors
                }, status=400)
            return render(request, 'finances/scheduled_transaction_form.html', {
                'form': form,
                'title': 'Create Scheduled Transaction'
            })
    else:
        # For AJAX requests asking for the form, return needed data
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'categories': [{'id': c.id, 'name': c.name, 'type': c.type} for c in Category.objects.filter(user=request.user)],
                'accounts': [{'id': a.id, 'name': a.name, 'type': getattr(a, 'get_account_type', lambda: 'unknown')()} 
                             for a in Account.objects.filter(user=request.user)]
            })
        
        form = ScheduledTransactionForm(user=request.user)
    return render(request, 'finances/scheduled_transaction_form.html', {
        'form': form,
        'title': 'Create Scheduled Transaction'
    })

@login_required
def edit_scheduled_transaction(request, pk):
    scheduled_transaction = get_object_or_404(ScheduledTransaction, pk=pk, user=request.user)

    if request.method == 'POST':
        form = ScheduledTransactionForm(request.POST, instance=scheduled_transaction, user=request.user)
        if form.is_valid():
            updated_transaction = form.save(commit=False)
            updated_transaction.is_recurring = updated_transaction.repeats == 0
            updated_transaction.save()
            
            # Check if it's an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Scheduled transaction updated successfully.',
                    'id': updated_transaction.id
                })
            
            messages.success(request, 'Scheduled transaction updated successfully!')
            return redirect('scheduled_transactions')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = {}
                for field, error_list in form.errors.items():
                    errors[field] = [str(error) for error in error_list]
                return JsonResponse({
                    'success': False,
                    'errors': errors
                })
            messages.error(request, 'There was an error with your input.')
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'id': scheduled_transaction.id,
                'name': scheduled_transaction.name,
                'amount': float(scheduled_transaction.amount),
                'transaction_type': scheduled_transaction.transaction_type,
                'date_scheduled': scheduled_transaction.date_scheduled.isoformat(),
                'repeat_type': scheduled_transaction.repeat_type,
                'repeats': scheduled_transaction.repeats,
                'note': scheduled_transaction.note or '',
                'category': scheduled_transaction.category_id if scheduled_transaction.category else None,
                'subcategory': scheduled_transaction.subcategory_id if hasattr(scheduled_transaction, 'subcategory') and scheduled_transaction.subcategory else None,
                'account': scheduled_transaction.account_id if scheduled_transaction.account else None
            })
        form = ScheduledTransactionForm(instance=scheduled_transaction, user=request.user)
    return render(request, 'finances/scheduled_transaction_form.html', {
        'form': form,
        'title': 'Edit Scheduled Transaction'
    })

@login_required
def delete_scheduled_transaction(request, pk):
    scheduled_transaction = get_object_or_404(ScheduledTransaction, pk=pk, user=request.user)

    related_count = ScheduledTransaction.objects.filter(parent_transaction=scheduled_transaction).count()
    if request.method == 'POST':
        if scheduled_transaction.parent_transaction is None:
            ScheduledTransaction.objects.filter(parent_transaction=scheduled_transaction).delete()
        scheduled_transaction.delete()
        messages.success(request, 'Scheduled transaction deleted successfully.')
        return redirect('scheduled_transactions')

    return render(request, 'finances/confirm_delete_scheduled_modal.html', {
        'object': scheduled_transaction,
        'related_count': related_count,
        'title': 'Delete Scheduled Transaction'
    })

@login_required
def debts_list(request):
    from datetime import date
    today = date.today()
    
    debts = Debt.objects.filter(user=request.user)

    debt_type = request.GET.get('type', '')
    search = request.GET.get('search', '')
    sort_by = request.GET.get('sort_by', 'date_issued')
    order = request.GET.get('order', 'asc')
    is_ajax = request.GET.get('ajax', 'false') == 'true'

    if debt_type:
        debts = debts.filter(debt_type=debt_type)
    
    if search:
        debts = debts.filter(
            Q(person__icontains=search) |
            Q(amount__icontains=search) |
            Q(notes__icontains=search)
        )

    account_filter = Q(user=request.user)
    
    debit_accounts = DebitAccount.objects.filter(account_filter)
    credit_accounts = CreditAccount.objects.filter(account_filter)
    wallet_accounts = Wallet.objects.filter(account_filter)
    
    # Combine the results in the context
    accounts = list(debit_accounts) + list(credit_accounts) + list(wallet_accounts)

    if sort_by == 'residual_amount':
        debts = debts.annotate(
            residual = ExpressionWrapper(
                F('amount') - F('paid'),
                output_field=DecimalField()
            )
        ).order_by('residual' if order == 'asc' else '-residual')
    else:
        if sort_by:
            debts = debts.order_by(sort_by)
            if order == 'desc':
                debts = debts.reverse()
    
    # Calculate summary totals
    total_debt_amount = sum(debt.amount for debt in debts if debt.debt_type == 'debt')
    total_credit_amount = sum(debt.amount for debt in debts if debt.debt_type == 'credit')
    total_debt_paid = sum(debt.paid for debt in debts if debt.debt_type == 'debt')
    net_position = total_credit_amount - total_debt_amount
    
    for debt in debts:
        # Calculate progress percentage
        debt.progress_percentage = (debt.paid / debt.amount) * 100 if debt.amount > 0 else 0
        
        # Calculate days until due
        debt.days_until_due = (debt.date_payback - today).days
    
    context = {
        'debts': debts,
        'accounts': accounts,
        'total_debt_amount': total_debt_amount,
        'total_credit_amount': total_credit_amount,
        'total_debt_paid': total_debt_paid,
        'net_position': net_position,
        'today': today,
    }

    # If it's an AJAX request, return just the content part
    if is_ajax:
        return render(request, 'finances/debts_content.html', context)
    
    # Otherwise, return the full page
    return render(request, 'finances/debts.html', context)

@login_required
def add_debt(request):
    if request.method == 'POST':
        form = DebtForm(request.POST)

        if form.is_valid():
            new_debt = form.save(commit=False)
            new_debt.user = request.user
            new_debt.save()

            messages.success(request, 'Debt Entry created successfully!')
            return redirect('debts_list')
        else:
            messages.error(request, 'There was an error with your input.')
    else:
        form = DebtForm()
        
    return render(request, 'finances/add_debt.html', {'form': form})

@login_required
def edit_debt(request, pk):
    debt = get_object_or_404(Debt, pk=pk)
    
    if request.method == "POST":
        form = DebtForm(request.POST, instance=debt)
        if form.is_valid():
            form.save()
            return redirect('debts_list')
    else:
        form = DebtForm(instance=debt)
    
    return render(request, 'finances/edit_debt.html', {'form': form, 'debt': debt})

@login_required
def delete_debt(request, pk):
    debt = get_object_or_404(Debt, pk=pk)
    
    if request.method == "POST":
        debt.delete()
        return redirect('debts_list')
    
    return render(request, 'finances/confirm_delete_debt.html', {'debt': debt})

def update_payment(request, pk):
    debt = get_object_or_404(Debt, pk=pk)
    
    if request.method == 'POST':
        paid_amount = Decimal(request.POST.get('paid_amount'))
        account_id = request.POST.get('account')
        
        # account = None
        account3 = Wallet.objects.filter(id=account_id).first()
        account1 = DebitAccount.objects.filter(id=account_id).first()
        account2 = CreditAccount.objects.filter(id=account_id).first()
        
        if paid_amount > debt.residual_amount:
            messages.error(request, "The payment amount exceeds the residual amount.")
            return redirect('debts_list')        
        
        if account1:
            if (account1.balance - paid_amount) < account1.maintaining_balance:
                messages.error(request, "Balance after deduction is below the maintaining balance.")
                return redirect('debts_list')
            # account = account1
            if debt.debt_type.lower() == 'credit':
                account1.balance += paid_amount
            else:
                account1.balance -= paid_amount
            account1.save()
        elif account2:
            if debt.debt_type.lower() == 'credit':
                messages.error(request, "Credit accounts cannot accept payments for credit-type debts.")
                return redirect('debts_list')
            if (account2.current_usage + paid_amount) > account2.credit_limit:
                messages.error(request, "Payment exceeds credit limit.")
                return redirect('debts_list')
            # account = account2
            account2.current_usage += paid_amount
            account2.save()
        elif account3:
            if account3.balance < paid_amount:
                messages.error(request, "Insufficient balance in wallet.")
                return redirect('debts_list')
            # account = account3
            if debt.debt_type.lower() == 'credit':
                account3.balance += paid_amount
            else:
                account3.balance -= paid_amount
            account3.save()        
        else:
            messages.error(request, "Account not found.")
            return redirect('debts_list')
        
        debt.paid += paid_amount
        debt.save()
            
        Transaction.objects.create(
            title="Debt Payment",
            amount=paid_amount,
            type="expense" if debt.debt_type == 'debt' else "income",
            category=None,
            subcategory=None,
            user=request.user,
            notes=f"Payment for debt with {debt.person}",
            # account=account
        )

        messages.success(request, "Payment successfully updated!")
        return redirect('debts_list')
    
    return redirect('debts_list')

@login_required
def accounts_list(request):
    debit_accounts = DebitAccount.objects.filter(user=request.user)
    credit_accounts = CreditAccount.objects.filter(user=request.user)
    wallet_accounts = Wallet.objects.filter(user=request.user)

    total_debit_balance = sum(account.balance for account in debit_accounts)
    total_credit_balance = sum(account.credit_limit - account.current_usage for account in credit_accounts)
    total_wallet_balance = sum(account.balance for account in wallet_accounts)
    total_balance = total_debit_balance + total_credit_balance + total_wallet_balance
    
    account_types = ['Debit', 'Credit', 'Wallet']
    account_names = []
    account_balances = []
    
    for account in debit_accounts:
        account_names.append(account.name)
        account_balances.append(account.balance)

    for account in credit_accounts:
        account_names.append(account.name)
        account_balances.append(account.credit_limit - account.current_usage)

    for account in wallet_accounts:
        account_names.append(account.name)
        account_balances.append(account.balance)
    
    total_account_type_balance = {
        'Debit': total_debit_balance,
        'Credit': total_credit_balance,
        'Wallet': total_wallet_balance,
    }

    total_balance_by_type = [
        total_debit_balance,
        total_credit_balance,
        total_wallet_balance,
    ]

    if request.method == 'POST':
        account_type = request.POST.get('account_type')

        form_classes = {
            'Debit': DebitAccountForm,
            'Credit': CreditAccountForm,
            'Wallet': WalletForm,
        }

        form_class = form_classes.get(account_type)
        form = form_class(request.POST)

        if form.is_valid():
            new_account = form.save(commit=False)
            new_account.user = request.user

            # Validation logic for Debit accounts
            if account_type == 'Debit':
                balance = form.cleaned_data.get('balance')
                maintaining_balance = form.cleaned_data.get('maintaining_balance')

                if maintaining_balance and balance < maintaining_balance:
                    messages.error(request, "Balance must be greater than or equal to the maintaining balance.")
                    return redirect('accounts_list')

            # Validation logic for Credit accounts
            if account_type == 'Credit':
                current_usage = form.cleaned_data.get('current_usage')
                credit_limit = form.cleaned_data.get('credit_limit')

                if current_usage > credit_limit:
                    messages.error(request, "Current usage cannot exceed the credit limit.")
                    return redirect('accounts_list')

            new_account.save()
            messages.success(request, f'{account_type} account created successfully!')
            return redirect('accounts_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = DebitAccountForm()
    
    chartdata = json.dumps({
        'total_balance': float(total_balance),
        'total_balance_by_type': [float(balance) for balance in total_balance_by_type],
        'account_names': account_names,
        'account_balances': [float(balance) for balance in account_balances],
    })

    context = {
        'account_types': account_types,
        'accounts': {
            'Debit': debit_accounts,
            'Credit': credit_accounts,
            'Wallet': wallet_accounts,
        },
        'chartdata': chartdata,
        'form': form,
    }

    return render(request, 'finances/accounts.html', context)

@login_required
def delete_account(request, account_id):
    account = get_object_or_404(Account, id=account_id, user=request.user)
    account.delete()
    return redirect('accounts_list')

def process_scheduled_transactions(user):
    now = timezone.now()
    
    due_transactions = ScheduledTransaction.objects.filter(
        user=user,
        date_scheduled__lte=now,
        status='scheduled'
    )
    
    for scheduled in due_transactions:
        try:
            # Validate account balance before processing
            account = scheduled.account
            amount_decimal = Decimal(str(scheduled.amount))
            
            if hasattr(account, 'debitaccount'):
                debit = account.debitaccount
                if scheduled.transaction_type == 'expense' and debit.balance - amount_decimal < (debit.maintaining_balance or 0):
                    scheduled.status = 'failed'
                    scheduled.save(update_fields=['status'])
                    continue
            elif hasattr(account, 'creditaccount'):
                credit = account.creditaccount
                if scheduled.transaction_type == 'expense' and credit.current_usage + amount_decimal > credit.credit_limit:
                    scheduled.status = 'failed'
                    scheduled.save(update_fields=['status'])
                    continue
            elif hasattr(account, 'wallet'):
                wallet = account.wallet
                if scheduled.transaction_type == 'expense' and wallet.balance < amount_decimal:
                    scheduled.status = 'failed'
                    scheduled.save(update_fields=['status'])
                    continue
            
            # Create the transaction
            transaction = Transaction.objects.create(
                user=user,
                title=scheduled.name,
                amount=scheduled.amount,
                date=scheduled.date_scheduled.date(),
                time=scheduled.date_scheduled.time(),
                type=scheduled.transaction_type,
                category=scheduled.category,
                subcategory=scheduled.subcategory,
                transaction_account=scheduled.account,
                notes=f"Scheduled transaction: {scheduled.name} (ID: {scheduled.id})"
            )
            
            # Update account balance/usage
            if hasattr(account, 'debitaccount'):
                debit = account.debitaccount
                if scheduled.transaction_type == 'expense':
                    debit.balance -= amount_decimal
                else:  # income
                    debit.balance += amount_decimal
                debit.save()
            elif hasattr(account, 'creditaccount'):
                credit = account.creditaccount
                if scheduled.transaction_type == 'expense':
                    credit.current_usage += amount_decimal
                else:  # income (rare)
                    credit.current_usage -= amount_decimal
                    if credit.current_usage < 0:
                        credit.current_usage = 0
                credit.save()
            elif hasattr(account, 'wallet'):
                wallet = account.wallet
                if scheduled.transaction_type == 'expense':
                    wallet.balance -= amount_decimal
                else:  # income
                    wallet.balance += amount_decimal
                wallet.save()
            
            # Update the scheduled transaction status to completed
            scheduled.status = 'completed'
            scheduled.save(update_fields=['status'])
            
            # If this is a recurring transaction, create the next occurrence
            if scheduled.repeats != 1:
                if scheduled.repeats == 0:
                    next_repeats = 0
                elif scheduled.repeats > 1:
                    next_repeats = scheduled.repeats - 1
                else:
                    next_repeats = None
                if next_repeats is not None:
                    next_date = scheduled.date_scheduled
                    if scheduled.repeat_type == 'daily':
                        next_date += timedelta(days=1)
                    elif scheduled.repeat_type == 'weekly':
                        next_date += timedelta(weeks=1)
                    elif scheduled.repeat_type == 'monthly':
                        if next_date.month == 12:
                            next_date = next_date.replace(year=next_date.year + 1, month=1)
                        else:
                            next_date = next_date.replace(month=next_date.month + 1)
                    elif scheduled.repeat_type == 'yearly':
                        next_date = next_date.replace(year=next_date.year + 1)
                    ScheduledTransaction.objects.create(
                        user=user,
                        name=scheduled.name,
                        category=scheduled.category,
                        subcategory=scheduled.subcategory,
                        transaction_type=scheduled.transaction_type,
                        account=scheduled.account,
                        amount=scheduled.amount,
                        date_scheduled=next_date,
                        repeat_type=scheduled.repeat_type,
                        repeats=next_repeats,
                        note=scheduled.note,
                        status='scheduled'
                    )
        except Exception as e:
            # If transaction creation fails, mark as failed using direct SQL update
            ScheduledTransaction.objects.filter(id=scheduled.id).update(status='failed')
            # Log the error
            logging.error(f"Failed to process scheduled transaction {scheduled.id}: {str(e)}")

@login_required
def dashboard(request):
    # Get user's active budgets
    budgets = Budget.objects.filter(
        user=request.user,
        end_date__gte=timezone.now().date()
    ).select_related('category', 'subcategory').order_by('end_date')

    # Separate weekly and monthly budgets
    weekly_budgets = []
    monthly_budgets = []
    for budget in budgets:
        budget.spent = budget.get_spent_amount()
        budget.percentage_used = (budget.spent / budget.amount * 100) if budget.amount > 0 else 0
        if budget.duration == '1 week':
            weekly_budgets.append(budget)
        elif budget.duration == '1 month':
            monthly_budgets.append(budget)

    process_scheduled_transactions(request.user)
    dashboard_preference, created = DashboardPreference.objects.get_or_create(user=request.user)

    current_month = timezone.now().date().replace(day=1)
    prev_month = (current_month - timedelta(days=1)).replace(day=1)
    next_month = (current_month + timedelta(days=32)).replace(day=1)

    # Account summaries
    accounts_summary = calculate_account_summaries(request.user)
    total_balance = Decimal('0.00')
    for account_type in accounts_summary.values():
        total_balance += account_type['balance']

    # Month summaries
    current_month_income, current_month_expenses = get_month_summary(request.user, current_month)
    current_month_balance = current_month_income - current_month_expenses

    prev_month_income, prev_month_expenses = get_month_summary(request.user, prev_month)
    prev_month_balance = prev_month_income - prev_month_expenses

    # Budget warnings
    budget_warnings = get_budget_warnings(request.user, current_month, current_month.replace(day=1, month=current_month.month+1) - timedelta(days=1))

    # Recent transactions
    recent_transactions = Transaction.objects.filter(user=request.user).order_by('-date', '-time')[:5]

    # Get upcoming scheduled transactions for the user
    upcoming_scheduled_transactions = ScheduledTransaction.objects.filter(
        user=request.user,
        status='scheduled',
        date_scheduled__gte=timezone.now()
    ).order_by('date_scheduled')[:5]  # Limit to 5 upcoming transactions
    
    # Get user's debts
    debts = Debt.objects.filter(
        user=request.user,
        paid=False
    ).order_by('date_payback')[:5]  # Limit to 5 unpaid debts
    
    # Get last 7 days data for balance chart
    today = timezone.now().date()
    week_ago = today - timezone.timedelta(days=7)
    
    last_7_days_data = {
        'labels': [],
        'income': [],
        'expenses': []
    }
    
    # Generate data for each day
    for i in range(7):
        current_date = week_ago + timezone.timedelta(days=i+1)
        last_7_days_data['labels'].append(current_date.strftime('%a'))
        
        # Get transactions for this day
        day_transactions = Transaction.objects.filter(user=request.user, date=current_date)
        
        # Calculate income and expenses for this day
        day_income = sum(t.amount for t in day_transactions if t.type == 'income')
        day_expenses = sum(t.amount for t in day_transactions if t.type == 'expense')
        
        last_7_days_data['income'].append(float(day_income))
        last_7_days_data['expenses'].append(float(day_expenses))
    
    # Get future projection data (next 30 days)
    future_balance_data = {
        'labels': [],
        'values': []
    }
    
    # Calculate starting balance
    current_balance = Decimal('0')
    for account in Account.objects.filter(user=request.user):
        if hasattr(account, 'debitaccount'):
            current_balance += account.debitaccount.balance
        elif hasattr(account, 'wallet'):
            current_balance += account.wallet.balance
        elif hasattr(account, 'creditaccount'):
            current_balance -= account.creditaccount.current_usage
    
    # Get scheduled transactions for next 30 days
    end_date = today + timezone.timedelta(days=30)
    scheduled_txns = ScheduledTransaction.objects.filter(
        user=request.user,
        status='scheduled',
        date_scheduled__gte=today,
        date_scheduled__lte=end_date
    ).order_by('date_scheduled')
    
    # Prepare projection
    projection_days = 30
    daily_amounts = [0] * projection_days
    
    # Add scheduled transactions to daily amounts
    for txn in scheduled_txns:
        day_index = (txn.date_scheduled.date() - today).days
        if 0 <= day_index < projection_days:
            if txn.transaction_type == 'income':
                daily_amounts[day_index] += float(txn.amount)
            else:
                daily_amounts[day_index] -= float(txn.amount)
    
    # Generate labels and calculate running balance
    balance = float(current_balance)
    for i in range(projection_days):
        future_date = today + timezone.timedelta(days=i)
        future_balance_data['labels'].append(future_date.strftime('%d %b'))
        
        balance += daily_amounts[i]
        future_balance_data['values'].append(balance)
    
    # Get category data for pie chart (last 30 days)
    month_ago = today - timezone.timedelta(days=30)
    categories_data = []
    
    # Get expense transactions for the last 30 days
    expense_transactions = Transaction.objects.filter(
        user=request.user,
        type='expense',
        date__gte=month_ago
    ).select_related('category')
    
    # Group transactions by category
    category_totals = {}
    for transaction in expense_transactions:
        category_name = transaction.category.name if transaction.category else 'Uncategorized'
        if category_name in category_totals:
            category_totals[category_name] += float(transaction.amount)
        else:
            category_totals[category_name] = float(transaction.amount)
    
    # Convert to list format for the chart
    for category, amount in category_totals.items():
        categories_data.append({
            'category': category,
            'amount': amount
        })
    
    # Sort by amount (descending)
    categories_data.sort(key=lambda x: x['amount'], reverse=True)
    
    # Limit to top 5 categories
    categories_data = categories_data[:5]

    context = {
        'accounts_summary': accounts_summary,
        'total_balance': total_balance,
        'current_month': current_month,
        'current_month_income': current_month_income,
        'current_month_expenses': current_month_expenses,
        'current_month_balance': current_month_balance,
        'prev_month': prev_month,
        'prev_month_income': prev_month_income,
        'prev_month_expenses': prev_month_expenses,
        'prev_month_balance': prev_month_balance,
        'budget_warnings': budget_warnings,
        'recent_transactions': recent_transactions,
        'dashboard_preference': dashboard_preference,
        'budgets': budgets,
        'weekly_budgets': weekly_budgets,
        'monthly_budgets': monthly_budgets,
        'upcoming_scheduled_transactions': upcoming_scheduled_transactions,
        'debts': debts,
        'last_7_days_data': json.dumps(last_7_days_data),
        'future_balance_data': json.dumps(future_balance_data),
        'categories_data': json.dumps(categories_data),
    }
    
    return render(request, 'finances/dashboard.html', context)

@login_required
def save_dashboard_preferences(request):
    """Save dashboard preferences to the database."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            dashboard_preference, created = DashboardPreference.objects.get_or_create(user=request.user)
            
            # Update dashboard preference
            if 'columns' in data:
                dashboard_preference.columns = data['columns']
            
            if 'visibleElements' in data:
                dashboard_preference.set_visible_elements(data['visibleElements'])
            
            if 'hiddenElements' in data:
                dashboard_preference.set_hidden_elements(data['hiddenElements'])
            
            dashboard_preference.save()
            
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    return JsonResponse({'status': 'error', 'message': 'Only POST method is allowed'}, status=405)

@login_required
def add_transaction(request):
    if request.method == 'POST':
        form = TransactionForm(request.POST, user=request.user)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.user = request.user
            transaction.save()
            messages.success(request, 'Transaction added successfully.')
            return redirect('dashboard')
    else:
        form = TransactionForm(user=request.user)
    
    return render(request, 'finances/transaction_form.html', {'form': form, 'title': 'Add Transaction'})

@login_required
def edit_transaction(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, user=request.user)
    
    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Transaction updated successfully.')
            return redirect('dashboard')
    else:
        form = TransactionForm(instance=transaction, user=request.user)
    
    return render(request, 'finances/transaction_form.html', {'form': form, 'title': 'Edit Transaction'})

@login_required
def delete_transaction(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk, user=request.user)
    
    if request.method == 'POST':
        transaction.delete()
        messages.success(request, 'Transaction deleted successfully.')
        return redirect('dashboard')
    
    return render(request, 'finances/transaction_confirm_delete.html', {'transaction': transaction})

@login_required
def reports_view(request):
    return render(request, 'reports.html')
    
@login_required
def monthly_summary(request):
    if request.method == 'POST':
        form = DateRangeForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
        else:
            today = timezone.now().date()
            start_date = today.replace(day=1)
            end_date = today
    else:
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
        form = DateRangeForm(initial={'start_date': start_date, 'end_date': end_date})
    
    # Get all transactions within the date range
    transactions = Transaction.objects.filter(
        user=request.user,
        date__gte=start_date,
        date__lte=end_date
    ).order_by('-date')
    
    # Calculate totals
    income = Transaction.objects.filter(
        user=request.user,
        type='income',
        date__gte=start_date,
        date__lte=end_date
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    expenses = Transaction.objects.filter(
        user=request.user,
        type='expense',
        date__gte=start_date,
        date__lte=end_date
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Get expenses by category for pie chart
    categories = Category.objects.filter(user=request.user)
    expenses_by_category = []
    
    for category in categories:
        amount = Transaction.objects.filter(
            user=request.user,
            type='expense',
            category=category,
            date__gte=start_date,
            date__lte=end_date
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        
        if amount > 0:
            expenses_by_category.append({
                'category': category.name,
                'amount': float(amount)
            })
    
    context = {
        'form': form,
        'transactions': transactions,
        'income': income,
        'expenses': expenses,
        'balance': income - expenses,
        'expenses_by_category': json.dumps(expenses_by_category),
        'start_date': start_date,
        'end_date': end_date
    }
    
    return render(request, 'finances/monthly_summary.html', context)

@login_required
def categories(request):
    """View to manage categories, works with both Django forms and React frontend"""
    
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.user = request.user
            category.save()
            
            # If AJAX request, return JSON response
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'category': {
                        'id': category.id,
                        'name': category.name,
                        'icon': category.icon
                    }
                })
            
            messages.success(request, 'Category added successfully.')
            return redirect('categories')
        else:
            # If AJAX request, return error response
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                }, status=400)
    else:
        form = CategoryForm(initial={'icon': 'bi-tag'})
    
    categories = Category.objects.filter(user=request.user)
    
    # Get available icon options for the form
    available_icons = [
        { 'icon': 'bi-tag', 'name': 'Tag' },
        { 'icon': 'bi-cash-coin', 'name': 'Cash' },
        { 'icon': 'bi-credit-card', 'name': 'Credit Card' },
        { 'icon': 'bi-cart', 'name': 'Shopping Cart' },
        { 'icon': 'bi-bag', 'name': 'Shopping Bag' },
        { 'icon': 'bi-truck', 'name': 'Transportation' },
        { 'icon': 'bi-house', 'name': 'Home' },
        { 'icon': 'bi-cup-hot', 'name': 'Food' },
        { 'icon': 'bi-music-note-beamed', 'name': 'Entertainment' },
        { 'icon': 'bi-people', 'name': 'Family' },
        { 'icon': 'bi-heart-pulse', 'name': 'Health' },
        { 'icon': 'bi-graph-up-arrow', 'name': 'Investments' },
        { 'icon': 'bi-gift', 'name': 'Gift' },
        { 'icon': 'bi-airplane', 'name': 'Travel' },
        { 'icon': 'bi-book', 'name': 'Education' },
        { 'icon': 'bi-lightning', 'name': 'Utilities' }
    ]
    
    context = {
        'form': form,
        'categories': categories,
        'available_icons': available_icons
    }
    
    return render(request, 'finances/categories.html', context)

@login_required
def export_csv(request):
    if request.method == 'POST':
        form = DateRangeForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            
            # Get all transactions within the date range
            transactions = Transaction.objects.filter(
                user=request.user,
                date__gte=start_date,
                date__lte=end_date
            ).order_by('-date')
            
            # Create CSV response
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="transactions_{start_date}_{end_date}.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Title', 'Amount', 'Date', 'Type', 'Category', 'Notes'])
            
            for transaction in transactions:
                category_name = transaction.category.name if transaction.category else 'Uncategorized'
                writer.writerow([
                    transaction.title,
                    transaction.amount,
                    transaction.date,
                    transaction.type,
                    category_name,
                    transaction.notes or ''
                ])
            
            return response
    else:
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
        form = DateRangeForm(initial={'start_date': start_date, 'end_date': end_date})
    
    return render(request, 'finances/export_csv.html', {'form': form})

@login_required
def manage_budget(request):
    form = BudgetForm(user=request.user)
    
    today = timezone.now().date()
    # Get all active budgets (those that include today's date)
    budgets = Budget.objects.select_related('subcategory', 'category').filter(
        user=request.user,
        start_date__lte=today,
        end_date__gte=today
    )
    
    budget_data = []
    has_weekly_budgets = False
    has_monthly_budgets = False
    
    for budget in budgets:
        # Check if we have any weekly or monthly budgets
        if budget.duration == '1 week':
            has_weekly_budgets = True
        elif budget.duration == '1 month':
            has_monthly_budgets = True
            
        # Calculate budget usage and add to data list
        budget_item = {
            'id': budget.id,
            'account': budget.account,
            'amount': budget.amount,
            'spent': budget.get_spent_amount(),
            'remaining': budget.get_remaining_amount(),
            'percentage_used': budget.get_percentage_spent(),
            'duration': budget.duration,
            'start_date': budget.start_date,
            'end_date': budget.end_date,
        }
        
        # Handle subcategory or category data
        if budget.subcategory:
            budget_item['subcategory'] = budget.subcategory
            budget_item['category'] = budget.subcategory.parent_category
        else:
            budget_item['subcategory'] = None
            budget_item['category'] = budget.category
            
        budget_data.append(budget_item)
    
    # Get all categories for the category selector - fix the filter to only get user's categories
    categories = Category.objects.filter(user=request.user)
    
    context = {
        'form': form,
        'budgets': budget_data,
        'categories': categories,
        'has_weekly_budgets': has_weekly_budgets,
        'has_monthly_budgets': has_monthly_budgets,
    }
    
    return render(request, 'finances/manage_budget.html', context)

@login_required
def update_budget(request):
    if request.method == 'POST':
        budget_id = request.POST['budget_id']  # Get the budget ID from the form
        budget = get_object_or_404(Budget, id=budget_id, user=request.user)

        # Get form data
        subcategory = SubCategory.objects.get(id=request.POST['subcategory'])
        amount = Decimal(request.POST['amount'])  # Convert amount to Decimal
        start_date = datetime.strptime(request.POST['start_date'], '%Y-%m-%d').date()  # Convert to date object
        duration = request.POST['duration']
        account = Account.objects.get(id=request.POST['account'])

        # Update the budget with new values
        budget.subcategory = subcategory
        budget.amount = amount
        budget.start_date = start_date
        budget.duration = duration
        budget.account = account
        
        # Calculate end date based on duration
        if duration == '1 week':
            budget.end_date = start_date + timedelta(days=6)
        else:  # 1 month
            if start_date.month == 12:
                next_month = start_date.replace(year=start_date.year + 1, month=1, day=1)
            else:
                next_month = start_date.replace(month=start_date.month + 1, day=1)
            budget.end_date = next_month - timedelta(days=1)

        budget.save()

        # Return the updated budget data as JSON
        updated_budget = {
            'subcategory': budget.subcategory.name,
            'amount': float(budget.amount),
            'spent': float(budget.spent),
            'remaining': float(budget.remaining),
            'percentage_used': budget.percentage_used,
            'start_date': budget.start_date.strftime('%Y-%m-%d'),
            'end_date': budget.end_date.strftime('%Y-%m-%d')
        }
        return JsonResponse({'success': True, 'updatedBudget': updated_budget})

    return JsonResponse({'success': False})

def get_budgets(request):
    budgets = Budget.objects.all().values(
        'subcategory', 'amount', 'spent', 'start_date', 'end_date', 'duration'
    )
    
    # Calculate 'remaining' dynamically for each budget
    budget_list = []
    for budget in budgets:
        budget['remaining'] = budget['amount'] - budget['spent']
        budget_list.append(budget)
    
    return JsonResponse(budget_list, safe=False)

def custom_logout(request):
    logout(request)
    messages.success(request, 'You have been successfully logged out.')
    return redirect('home')

@login_required
def add_budget(request):
    if request.method == 'POST':
        form = BudgetForm(request.POST, user=request.user)
        if form.is_valid():
            budget = form.save(commit=False)
            budget.user = request.user
            
            # When no subcategory is provided, use the category directly
            if not budget.category and 'category' in request.POST:
                try:
                    category_id = request.POST.get('category')
                    budget.category = Category.objects.get(id=category_id, user=request.user)
                except (Category.DoesNotExist, ValueError):
                    return JsonResponse({
                        'success': False, 
                        'errors': {
                            'category': ['Please select a valid category.']
                        }
                    })
            
            # Check for existing budget in the same period
            if budget.category:
                existing_budget = Budget.objects.filter(
                    user=request.user,
                    category=budget.category,
                    start_date__year=budget.start_date.year,
                    start_date__month=budget.start_date.month,
                    account=budget.account
                ).first()
                
                if existing_budget:
                    return JsonResponse({
                        'success': False, 
                        'errors': {
                            'category': ['A budget for this category already exists in the selected period.']
                        }
                    })
                
                budget.save()
                return JsonResponse({'success': True})
            else:
                return JsonResponse({
                    'success': False, 
                    'errors': {
                        'subcategory': ['A budget for this subcategory already exists in the selected period.']
                    }
                })
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def delete_budget(request):
    if request.method == 'POST':
        budget_id = request.POST.get('budget_id')
        try:
            budget = Budget.objects.get(id=budget_id, user=request.user)
            budget.delete()
            return JsonResponse({'success': True})
        except Budget.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Budget not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def categories_chart(request):
    # Placeholder data - will be replaced with actual data later
    context = {
        'chart_data': {
            'labels': ['Food', 'Transport', 'Entertainment', 'Bills', 'Shopping'],
            'data': [30, 20, 15, 25, 10],
            'colors': ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF']
        }
    }
    return render(request, 'finances/categories_chart.html', context)

@login_required
def future_projections(request):
    # Placeholder data - will be replaced with actual data later
    context = {
        'projections': {
            'labels': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'],
            'income': [2000, 2200, 2100, 2300, 2400, 2500],
            'expenses': [1800, 1900, 2000, 2100, 2200, 2300],
            'savings': [200, 300, 100, 200, 200, 200]
        }
    }
    return render(request, 'finances/future_projections.html', context)

@login_required
def time_analysis(request):
    # Placeholder data - will be replaced with actual data later
    context = {
        'time_data': {
            'labels': ['Week 1', 'Week 2', 'Week 3', 'Week 4'],
            'spending': [500, 600, 450, 700],
            'income': [1000, 1000, 1000, 1000]
        }
    }
    return render(request, 'finances/time_analysis.html', context)

def custom_logout(request):
    logout(request)
    messages.success(request, 'You have been successfully logged out.')
    return redirect('home')

@login_required
def api_categories(request):
    """API endpoint to get categories"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    logger = logging.getLogger(__name__)
    
    # Get all categories for the user, ordered by the order field
    categories = Category.objects.filter(user=request.user).order_by('order', 'name')
    logger.info(f"User {request.user.username} has {categories.count()} categories")
    
    income_categories = []
    expense_categories = []
    
    for category in categories:
        try:
            # Get subcategories for each category with explicit query to ensure they're loaded
            subcategories = list(SubCategory.objects.filter(parent_category=category).values('id', 'name', 'icon'))
            logger.info(f"Category {category.name} has {len(subcategories)} subcategories")
            
            category_data = {
                'id': category.id,
                'name': category.name,
                'icon': category.icon,
                'type': category.type,  # Include type in the response
                'order': category.order,  # Include order in the response
                'subcategories': subcategories
            }
            
            # Classify based on the type field
            if category.type == 'income':
                income_categories.append(category_data)
            else:
                expense_categories.append(category_data)
                
            logger.info(f"Categorized {category.name} as {category.type}")
        except Exception as e:
            # Log the error but continue with other categories
            logger.error(f"Error processing category {category.name}: {str(e)}")
            continue
    
    # Get list of Bootstrap icons
    available_icons = get_available_icons()
    
    response_data = {
        'income': income_categories,
        'expenses': expense_categories,
        'availableIcons': available_icons
    }
    
    # Log sample of the response for debugging
    if income_categories:
        logger.info(f"Sample income category: {income_categories[0]['name']} with {len(income_categories[0]['subcategories'])} subcategories")
    if expense_categories:
        logger.info(f"Sample expense category: {expense_categories[0]['name']} with {len(expense_categories[0]['subcategories'])} subcategories")
    
    return JsonResponse(response_data)

def api_add_category(request):
    """API endpoint to add a category"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST request required'}, status=400)
    
    # Debug output to see what's being received
    print(f"Adding category with POST data: {request.POST}")
    print(f"Type field value: {request.POST.get('type', 'not provided')}")
    
    form = CategoryForm(request.POST)
    if form.is_valid():
        category = form.save(commit=False)
        category.user = request.user
        
        # Explicitly set the type based on the POST data
        if 'type' in request.POST:
            category_type = request.POST['type']
            if category_type == 'income':
                category.type = 'income'
                print("Setting category type to income")
            elif category_type == 'expense':
                category.type = 'expense'
                print("Setting category type to expense")
            else:
                print(f"Unrecognized type value: {category_type}, defaulting to expense")
                category.type = 'expense'
        else:
            print("No type provided, defaulting to expense")
            category.type = 'expense'
            
        category.save()
        
        print(f"Category saved with type: {category.type}")
        
        return JsonResponse({
            'success': True,
            'category': {
                'id': category.id,
                'name': category.name,
                'icon': category.icon,
                'type': category.type,  # Include type in response
                'subcategories': []
            }
        })
    else:
        print(f"Form validation errors: {form.errors}")
        return JsonResponse({'error': 'Invalid form data', 'errors': form.errors}, status=400)
    
def api_subcategories(request, category_id):
    """API endpoint to get subcategories for a category"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    try:
        category = Category.objects.get(id=category_id, user=request.user)
        subcategories = SubCategory.objects.filter(parent_category=category)
        
        # Debug info
        print(f"User {request.user.username} requested subcategories for category {category.name} (ID: {category_id})")
        print(f"Found {subcategories.count()} subcategories")
        
        subcategories_data = []
        for subcategory in subcategories:
            subcategories_data.append({
                'id': subcategory.id,
                'name': subcategory.name,
                'icon': subcategory.icon
            })
            print(f"  - Subcategory: {subcategory.name} (ID: {subcategory.id})")
        
        return JsonResponse(subcategories_data, safe=False)
    except Category.DoesNotExist:
        print(f"Category with ID {category_id} not found for user {request.user.username}")
        return JsonResponse({'error': 'Category not found'}, status=404)
    
# API endpoint to get a single subcategory details
def api_subcategory_detail(request, subcategory_id):
    """API endpoint to get details of a specific subcategory"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    try:
        subcategory = SubCategory.objects.get(id=subcategory_id, parent_category__user=request.user)
        
        subcategory_data = {
            'id': subcategory.id,
            'name': subcategory.name,
            'icon': subcategory.icon,
            'parent_category': subcategory.parent_category.id
        }
        
        return JsonResponse(subcategory_data)
    except SubCategory.DoesNotExist:
        return JsonResponse({'error': 'Subcategory not found'}, status=404)

# API endpoint for adding a subcategory
def api_add_subcategory(request, category_id):
    """API endpoint to add a subcategory to a category"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST request required'}, status=400)
    
    try:
        category = Category.objects.get(id=category_id, user=request.user)
    except Category.DoesNotExist:
        return JsonResponse({'error': 'Category not found'}, status=404)
    
    form = SubCategoryForm(request.POST, user=request.user, category_id=category_id)
    if form.is_valid():
        subcategory = form.save(commit=False)
        subcategory.parent_category = category
        subcategory.save()

        # Check if the category already has itself as a subcategory
        category_as_subcategory = SubCategory.objects.filter(
            parent_category=category,
            name=category.name,
        ).first()
        
        # If the category doesn't have itself as a subcategory, create it
        if not category_as_subcategory:
            category_as_subcategory = SubCategory.objects.create(
                name=category.name,
                parent_category=category,
                icon=category.icon
            )
        
        return JsonResponse({
            'success': True,
            'subcategory': {
                'id': subcategory.id,
                'name': subcategory.name,
                'icon': subcategory.icon
            }
        })
    else:
        return JsonResponse({'error': 'Invalid form data', 'errors': form.errors}, status=400)

# API endpoint for deleting a subcategory
def api_delete_subcategory(request, subcategory_id):
    """API endpoint to delete a subcategory"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    if request.method != 'DELETE':
        return JsonResponse({'error': 'DELETE request required'}, status=400)
    
    try:
        subcategory = SubCategory.objects.get(id=subcategory_id, parent_category__user=request.user)
        subcategory.delete()
        return JsonResponse({'success': True})
    except SubCategory.DoesNotExist:
        return JsonResponse({'error': 'Subcategory not found'}, status=404)

# API endpoint for updating a subcategory
def api_update_subcategory(request, subcategory_id):
    """API endpoint to update a subcategory"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST request required'}, status=400)
    
    try:
        subcategory = SubCategory.objects.get(id=subcategory_id, parent_category__user=request.user)
        
        # Update subcategory fields
        name = request.POST.get('name', '').strip()
        icon = request.POST.get('icon', subcategory.icon)
        
        if not name:
            return JsonResponse({'error': 'Subcategory name cannot be empty'}, status=400)
        
        subcategory.name = name
        subcategory.icon = icon
        subcategory.save()
        
        return JsonResponse({
            'success': True,
            'subcategory': {
                'id': subcategory.id,
                'name': subcategory.name,
                'icon': subcategory.icon
            }
        })
    except SubCategory.DoesNotExist:
        return JsonResponse({'error': 'Subcategory not found'}, status=404)

# API endpoint for deleting a category
def api_delete_category(request, category_id):
    """API endpoint to delete a category"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    if request.method != 'DELETE':
        return JsonResponse({'error': 'DELETE request required'}, status=400)
    
    try:
        category = Category.objects.get(id=category_id, user=request.user)
        
        # Delete the category (this will also delete all subcategories due to CASCADE)
        category.delete()
        return JsonResponse({'success': True})
    except Category.DoesNotExist:
        return JsonResponse({'error': 'Category not found'}, status=404)

# API endpoint for updating a category
def api_update_category(request, category_id):
    """API endpoint to update a category"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST request required'}, status=400)
    
    try:
        category = Category.objects.get(id=category_id, user=request.user)
        
        # Update category fields
        name = request.POST.get('name', '').strip()
        icon = request.POST.get('icon', category.icon)
        
        if not name:
            return JsonResponse({'error': 'Category name cannot be empty'}, status=400)
        
        # Check if the name is already used by another category
        if Category.objects.filter(user=request.user, name=name).exclude(id=category_id).exists():
            return JsonResponse({'error': 'A category with this name already exists'}, status=400)
        
        category.name = name
        category.icon = icon
        category.save()
        
        return JsonResponse({
            'success': True,
            'category': {
                'id': category.id,
                'name': category.name,
                'icon': category.icon,
                'type': category.type
            }
        })
    except Category.DoesNotExist:
        return JsonResponse({'error': 'Category not found'}, status=404)

def get_icon_for_category(category_name):
    """Get a suitable icon for a category based on its name"""
    category_name = category_name.lower()
    
    if any(word in category_name for word in ['salary', 'income', 'wage']):
        return 'bi-cash-coin'
    elif any(word in category_name for word in ['invest', 'stock', 'dividend', 'return']):
        return 'bi-graph-up-arrow'
    elif any(word in category_name for word in ['food', 'grocery', 'restaurant', 'dining']):
        return 'bi-cup-hot'
    elif any(word in category_name for word in ['transport', 'fuel', 'car', 'bus', 'train']):
        return 'bi-truck'
    elif any(word in category_name for word in ['house', 'rent', 'mortgage', 'property']):
        return 'bi-house'
    elif any(word in category_name for word in ['health', 'medical', 'doctor', 'medicine']):
        return 'bi-heart-pulse'
    elif any(word in category_name for word in ['gift', 'present', 'donation']):
        return 'bi-gift'
    elif any(word in category_name for word in ['shop', 'purchase', 'buy']):
        return 'bi-cart'
    elif any(word in category_name for word in ['bill', 'utility', 'electric', 'water', 'gas']):
        return 'bi-lightning'
    elif any(word in category_name for word in ['education', 'school', 'learn', 'course', 'book']):
        return 'bi-book'
    elif any(word in category_name for word in ['travel', 'holiday', 'vacation', 'trip']):
        return 'bi-airplane'
    elif any(word in category_name for word in ['entertainment', 'movie', 'game', 'fun']):
        return 'bi-music-note-beamed'
    
    # Default icon
    return 'bi-tag'

def get_available_icons():
    """Return list of available Bootstrap icons for categories"""
    return [
        { 'icon': 'bi-tag', 'name': 'Tag' },
        { 'icon': 'bi-tag-fill', 'name': 'Tag (Filled)' },
        { 'icon': 'bi-cash-coin', 'name': 'Cash' },
        { 'icon': 'bi-credit-card', 'name': 'Credit Card' },
        { 'icon': 'bi-cart', 'name': 'Shopping Cart' },
        { 'icon': 'bi-bag', 'name': 'Shopping Bag' },
        { 'icon': 'bi-truck', 'name': 'Transportation' },
        { 'icon': 'bi-house', 'name': 'Home' },
        { 'icon': 'bi-cup-hot', 'name': 'Food' },
        { 'icon': 'bi-music-note-beamed', 'name': 'Entertainment' },
        { 'icon': 'bi-people', 'name': 'Family' },
        { 'icon': 'bi-heart-pulse', 'name': 'Health' },
        { 'icon': 'bi-graph-up-arrow', 'name': 'Investments' },
        { 'icon': 'bi-gift', 'name': 'Gift' },
        { 'icon': 'bi-airplane', 'name': 'Travel' },
        { 'icon': 'bi-book', 'name': 'Education' },
        { 'icon': 'bi-lightning', 'name': 'Utilities' }
    ]

@login_required
def api_reorder_categories(request):
    """API endpoint to update the order of categories"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Parse JSON data
        data = json.loads(request.body)
        category_type = data.get('category_type')
        category_ids = data.get('category_ids', [])
        
        # Validate data
        if not category_type or not category_ids:
            return JsonResponse({'error': 'Missing required data'}, status=400)
        
        if category_type not in ['income', 'expense']:
            return JsonResponse({'error': 'Invalid category type'}, status=400)
        
        # Check if all category ids belong to the user
        categories = Category.objects.filter(
            user=request.user, 
            id__in=category_ids,
            type=category_type
        )
        
        # Check if we found all the categories
        if categories.count() != len(category_ids):
            return JsonResponse({'error': 'One or more categories not found'}, status=404)
        
        # Update the order for each category
        for index, category_id in enumerate(category_ids):
            category = categories.get(id=category_id)
            category.order = index
            category.save()
        
        return JsonResponse({'success': True})
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def transactions(request):
    """View for the transactions page with filtering and monthly view"""
    
    # Get current month and year, defaulting to current date
    current_date = timezone.now().date()
    current_month = int(request.GET.get('month', current_date.month))
    current_year = int(request.GET.get('year', current_date.year))
    
    # Get start and end date for the selected month
    start_date = timezone.datetime(current_year, current_month, 1).date()
    if current_month == 12:
        end_date = timezone.datetime(current_year + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = timezone.datetime(current_year, current_month + 1, 1).date() - timedelta(days=1)
    
    # Get filter parameters
    category_id = request.GET.get('category', '')
    subcategory_id = request.GET.get('subcategory', '')
    search_term = request.GET.get('search', '')
    types_param = request.GET.get('types', 'expense,income')
    types = types_param.split(',') if types_param else ['expense', 'income']
    
    # Start with base query for the current month
    transactions_query = Transaction.objects.filter(
        user=request.user,
        date__gte=start_date,
        date__lte=end_date
    )
    
    # Apply filters if present
    if category_id:
        transactions_query = transactions_query.filter(category_id=category_id)
        
        if subcategory_id:
            transactions_query = transactions_query.filter(subcategory_id=subcategory_id)
    
    if types and 'all' not in types:
        transactions_query = transactions_query.filter(type__in=types)
    
    if search_term:
        transactions_query = transactions_query.filter(
            models.Q(title__icontains=search_term) | 
            models.Q(notes__icontains=search_term)
        )
    
    # Get all transactions for the month with filters applied
    transactions = transactions_query.order_by('-date')
    
    # Calculate total
    income = Transaction.objects.filter(
        user=request.user,
        type='income',
        date__gte=start_date,
        date__lte=end_date
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    expenses = Transaction.objects.filter(
        user=request.user,
        type='expense',
        date__gte=start_date,
        date__lte=end_date
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    total_balance = income - expenses
    
    # Get all categories for filter and forms
    categories = Category.objects.filter(user=request.user).order_by('order', 'name')
    
    # Get all accounts
    accounts = Account.objects.filter(user=request.user)
    
    # Format month name
    month_names = ['January', 'February', 'March', 'April', 'May', 'June', 
                'July', 'August', 'September', 'October', 'November', 'December']
    current_month_name = month_names[current_month - 1]
    
    context = {
        'transactions': transactions,
        'categories': categories,
        'accounts': accounts,  # Add accounts to context
        'current_month': current_month,
        'current_year': current_year,
        'current_month_name': current_month_name,
        'total_balance': total_balance,
        'total_income': income,
        'total_expenses': expenses,
        'selected_category': category_id,
        'selected_subcategory': subcategory_id,
    }
    
    return render(request, 'finances/transactions.html', context)

@login_required
def transactions_api(request):
    """API endpoint for fetching transactions with filters"""
    
    # Get parameters from request
    month_param = request.GET.get('month')
    if month_param and '-' in month_param:
        try:
            year, month = map(int, month_param.split('-'))
            current_month = month
            current_year = year
        except (ValueError, IndexError):
            current_month = timezone.now().date().month
            current_year = timezone.now().date().year
    else:
        current_month = int(request.GET.get('month', timezone.now().date().month))
        current_year = int(request.GET.get('year', timezone.now().date().year))
    
    category_id = request.GET.get('category', '')
    subcategory_id = request.GET.get('subcategory', '')
    search_term = request.GET.get('search', '')
    types_param = request.GET.get('types', 'expense,income')
    types = types_param.split(',') if types_param else ['expense', 'income']
    
    # Get start and end date for the selected month
    start_date = timezone.datetime(current_year, current_month, 1).date()
    if current_month == 12:
        end_date = timezone.datetime(current_year + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = timezone.datetime(current_year, current_month + 1, 1).date() - timedelta(days=1)
    
    # Filter transactions and use select_related to fetch category and subcategory in a single query
    transactions = Transaction.objects.filter(
        user=request.user,
        date__gte=start_date,
        date__lte=end_date
    ).select_related('category', 'subcategory', 'transaction_account')
    
    # Apply type filter if specified
    if types and 'all' not in types:
        transactions = transactions.filter(type__in=types)
    
    # Additional filters
    if category_id:
        transactions = transactions.filter(category_id=category_id)
        
        # If subcategory is specified, filter by it as well
        if subcategory_id:
            transactions = transactions.filter(subcategory_id=subcategory_id)
    
    if search_term:
        transactions = transactions.filter(
            models.Q(title__icontains=search_term) | 
            models.Q(notes__icontains=search_term)
        )
    
    # Order by date (newest first)
    transactions = transactions.order_by('-date')
    
    # Calculate totals
    income = 0
    expenses = 0
    for t in transactions:
        try:
            if t.type == 'income':
                income += float(t.amount)
            elif t.type == 'expense':
                expenses += float(t.amount)
        except (ValueError, decimal.InvalidOperation):
            # Skip transactions with invalid amount values
            continue
    
    total = income - expenses
    
    # Format month name for response
    month_names = ['January', 'February', 'March', 'April', 'May', 'June', 
                  'July', 'August', 'September', 'October', 'November', 'December']
    current_month_name = month_names[current_month - 1]
    
    # Prepare data for JSON response
    transactions_data = []
    for transaction in transactions:
        # Get category information
        category_name = transaction.category.name if transaction.category else None
        category_icon = transaction.category.icon if transaction.category else 'bi bi-tag'
        
        # Get subcategory information
        subcategory_name = None
        subcategory_icon = None
        if transaction.subcategory:
            subcategory_name = transaction.subcategory.name
            subcategory_icon = transaction.subcategory.icon or category_icon
        
        # Get account information
        account_name = transaction.transaction_account.name if transaction.transaction_account else None
        
        # Use subcategory info if it exists, otherwise use category info
        display_name = subcategory_name or category_name or 'Uncategorized'
        display_icon = subcategory_icon or category_icon or 'bi bi-tag'
        
        # Format time if available
        time_str = None
        if hasattr(transaction, 'time') and transaction.time:
            time_str = transaction.time.strftime('%H:%M')
        
        transactions_data.append({
            'id': transaction.id,
            'title': transaction.title,
            'amount': float(transaction.amount),
            'date': transaction.date.isoformat(),
            'time': time_str,
            'type': transaction.type,
            'category': transaction.category_id if transaction.category else None,
            'subcategory': transaction.subcategory_id if transaction.subcategory else None,
            'transaction_account': transaction.transaction_account_id if transaction.transaction_account else None,
            'category_name': category_name,
            'category_icon': category_icon,
            'subcategory_name': subcategory_name,
            'subcategory_icon': subcategory_icon,
            'account_name': account_name,
            'display_name': display_name,
            'display_icon': display_icon,
            'notes': transaction.notes or '',
        })
    
    # Return complete data
    return JsonResponse({
        'transactions': transactions_data,
        'total': total,
        'income': income,
        'expenses': expenses,
        'current_month': current_month,
        'current_year': current_year,
        'current_month_name': current_month_name,
        'filters': {
            'category': category_id,
            'subcategory': subcategory_id,
            'search': search_term,
            'types': types
        }
    })

@login_required
def transaction_detail_api(request, transaction_id):
    """API endpoint for getting a single transaction's details"""
    print(f"Fetching details for transaction ID: {transaction_id}")
    
    try:
        transaction = get_object_or_404(Transaction, id=transaction_id, user=request.user)
        print(f"Transaction found: {transaction}")
        
        # Load the related objects to ensure they're available
        if transaction.category:
            category_name = transaction.category.name
            category_icon = transaction.category.icon
        else:
            category_name = None
            category_icon = 'bi bi-tag'
            
        # Load subcategory info if available
        if transaction.subcategory:
            subcategory_name = transaction.subcategory.name
            subcategory_icon = transaction.subcategory.icon or category_icon
        else:
            subcategory_name = None
            subcategory_icon = None
        
            # Format time if available
            time_str = None
            if hasattr(transaction, 'time') and transaction.time:
                time_str = transaction.time.strftime('%H:%M')
            
        data = {
            'id': transaction.id,
            'title': transaction.title,
            'amount': float(transaction.amount),
            'date': transaction.date.isoformat(),
                'time': time_str,
            'type': transaction.type,
            'category': transaction.category_id if transaction.category else None,
            'subcategory': transaction.subcategory_id if transaction.subcategory else None,
            'transaction_account': transaction.transaction_account_id if transaction.transaction_account else None,
            'category_name': category_name,
            'category_icon': category_icon,
            'subcategory_name': subcategory_name,
            'subcategory_icon': subcategory_icon,
            'notes': transaction.notes or '',
        }
        
        print(f"Returning transaction data: {data}")
        return JsonResponse(data)
    except Exception as e:
        print(f"Error fetching transaction details: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def create_transaction_api(request):
    """API endpoint for creating a transaction"""
    if request.method == 'POST':
        try:
            # Debug: Print out all POST data
            print("Creating transaction with POST data:", request.POST)
            
            title = request.POST.get('title')
            amount = request.POST.get('amount')
            date_str = request.POST.get('date')
            transaction_type = request.POST.get('type')
            category_id = request.POST.get('category') or None
            subcategory_id = request.POST.get('subcategory') or None
            transaction_account_id = request.POST.get('transaction_account') or None
            notes = request.POST.get('notes', '')
            
            # Debug: Print all parameters
            print(f"Title: {title}, Amount: {amount}, Date: {date_str}, Type: {transaction_type}")
            print(f"Category ID: {category_id}, Subcategory ID: {subcategory_id}, Notes: {notes}")
            
            # Validate required fields
            if not title or not amount or not date_str or not transaction_type:
                return JsonResponse({'success': False, 'error': 'Missing required fields'})
            
            # Convert amount to float
            try:
                amount = float(amount.replace(',', '.').strip())
            except (ValueError, AttributeError, decimal.InvalidOperation):
                return JsonResponse({'success': False, 'error': 'Invalid amount format'})
            
            # Create the transaction
            transaction = Transaction(
                title=title,
                amount=amount,
                date=date_str,
                type=transaction_type,
                user=request.user,
                notes=notes
            )
            
            if category_id and category_id != 'null' and category_id != '':
                try:
                    category = get_object_or_404(Category, id=category_id, user=request.user)
                    transaction.category = category
                except Exception as e:
                    print(f"Error getting category: {e}")
            
            if subcategory_id and subcategory_id != 'null' and subcategory_id != '':
                try:
                    subcategory = get_object_or_404(SubCategory, id=subcategory_id, parent_category__user=request.user)
                    transaction.subcategory = subcategory
                except Exception as e:
                    print(f"Error getting subcategory: {e}")
            
            if transaction_account_id and transaction_account_id != 'null' and transaction_account_id != '':
                try:
                    account = get_object_or_404(Account, id=transaction_account_id, user=request.user)
                    transaction.transaction_account = account
                except Exception as e:
                    print(f"Error getting account: {e}")
            
            transaction.save()
            print(f"Transaction created successfully with ID: {transaction.id}")

            # --- Update account balance/usage if transaction_account is set ---
            if transaction.transaction_account:
                account = transaction.transaction_account
                amount_decimal = Decimal(str(transaction.amount))
                # Debit Account
                if hasattr(account, 'debitaccount'):
                    debit = account.debitaccount
                    if transaction.type == 'expense':
                        debit.balance -= amount_decimal
                    else:  # income
                        debit.balance += amount_decimal
                    debit.save()
                # Credit Account
                elif hasattr(account, 'creditaccount'):
                    credit = account.creditaccount
                    if transaction.type == 'expense':
                        credit.current_usage += amount_decimal
                    else:  # income (rare)
                        credit.current_usage -= amount_decimal
                        if credit.current_usage < 0:
                            credit.current_usage = 0
                    credit.save()
                # Wallet
                elif hasattr(account, 'wallet'):
                    wallet = account.wallet
                    if transaction.type == 'expense':
                        wallet.balance -= amount_decimal
                    else:  # income
                        wallet.balance += amount_decimal
                    wallet.save()
            # --- End account update logic ---

            return JsonResponse({'success': True, 'transaction_id': transaction.id})
        except Exception as e:
            print(f"Error creating transaction: {e}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def update_transaction_api(request, transaction_id):
    """API endpoint for updating a transaction"""
    transaction = get_object_or_404(Transaction, id=transaction_id, user=request.user)
    
    if request.method == 'POST':
        try:
            # Debug: Print out all POST data
            print(f"Updating transaction {transaction_id} with POST data:", request.POST)
            
            title = request.POST.get('title')
            amount = request.POST.get('amount')
            date_str = request.POST.get('date')
            transaction_type = request.POST.get('type')
            category_id = request.POST.get('category') or None
            subcategory_id = request.POST.get('subcategory') or None
            transaction_account_id = request.POST.get('transaction_account') or None
            notes = request.POST.get('notes', '')
            
            # Debug: Print all parameters
            print(f"Title: {title}, Amount: {amount}, Date: {date_str}, Type: {transaction_type}")
            print(f"Category ID: {category_id}, Subcategory ID: {subcategory_id}, Notes: {notes}")
            print(f"Transaction Account ID: {transaction_account_id}")
            
            # Validate required fields
            if not title or not amount or not date_str or not transaction_type:
                return JsonResponse({'success': False, 'error': 'Missing required fields'})
            
            # Convert amount to float
            try:
                amount = float(amount.replace(',', '.').strip())
            except (ValueError, AttributeError, decimal.InvalidOperation):
                return JsonResponse({'success': False, 'error': 'Invalid amount format'})
            
            # Update the transaction
            transaction.title = title
            transaction.amount = amount
            transaction.date = date_str
            transaction.type = transaction_type
            transaction.notes = notes
            
            # Update category
            if category_id and category_id != 'null' and category_id != '':
                try:
                    category = get_object_or_404(Category, id=category_id, user=request.user)
                    transaction.category = category
                except Exception as e:
                    print(f"Error updating category: {e}")
            else:
                transaction.category = None
            
            # Update subcategory
            if subcategory_id and subcategory_id != 'null' and subcategory_id != '':
                try:
                    subcategory = get_object_or_404(SubCategory, id=subcategory_id, parent_category__user=request.user)
                    transaction.subcategory = subcategory
                except Exception as e:
                    print(f"Error updating subcategory: {e}")
            else:
                transaction.subcategory = None
            
            # Update account
            if transaction_account_id and transaction_account_id != 'null' and transaction_account_id != '':
                try:
                    account = get_object_or_404(Account, id=transaction_account_id, user=request.user)
                    transaction.transaction_account = account
                except Exception as e:
                    print(f"Error updating account: {e}")
            else:
                transaction.transaction_account = None
            
            transaction.save()
            print(f"Transaction updated successfully with ID: {transaction.id}")
            
            # --- Update account balance/usage after updating transaction ---
            # First, revert the previous impact on the account
            old_transaction = get_object_or_404(Transaction, id=transaction_id, user=request.user)
            old_account = old_transaction.transaction_account
            old_amount = Decimal(str(old_transaction.amount))
            
            if old_account:
                # Revert the previous transaction's effect
                if hasattr(old_account, 'debitaccount'):
                    debit = old_account.debitaccount
                    if old_transaction.type == 'expense':
                        debit.balance += old_amount  # Add back the expense amount
                    else:  # income
                        debit.balance -= old_amount  # Subtract the income amount
                    debit.save()
                elif hasattr(old_account, 'creditaccount'):
                    credit = old_account.creditaccount
                    if old_transaction.type == 'expense':
                        credit.current_usage -= old_amount  # Reduce the usage
                    else:  # income
                        credit.current_usage += old_amount  # Increase the usage
                        if credit.current_usage > credit.credit_limit:
                            credit.current_usage = credit.credit_limit
                    credit.save()
                elif hasattr(old_account, 'wallet'):
                    wallet = old_account.wallet
                    if old_transaction.type == 'expense':
                        wallet.balance += old_amount  # Add back the expense amount
                    else:  # income
                        wallet.balance -= old_amount  # Subtract the income amount
                    wallet.save()
            
            # Now apply the updated transaction to the selected account
            if transaction.transaction_account:
                account = transaction.transaction_account
                amount_decimal = Decimal(str(transaction.amount))
                
                # Debit Account
                if hasattr(account, 'debitaccount'):
                    debit = account.debitaccount
                    if transaction.type == 'expense':
                        debit.balance -= amount_decimal
                    else:  # income
                        debit.balance += amount_decimal
                    debit.save()
                # Credit Account
                elif hasattr(account, 'creditaccount'):
                    credit = account.creditaccount
                    if transaction.type == 'expense':
                        credit.current_usage += amount_decimal
                    else:  # income (rare)
                        credit.current_usage -= amount_decimal
                        if credit.current_usage < 0:
                            credit.current_usage = 0
                    credit.save()
                # Wallet
                elif hasattr(account, 'wallet'):
                    wallet = account.wallet
                    if transaction.type == 'expense':
                        wallet.balance -= amount_decimal
                    else:  # income
                        wallet.balance += amount_decimal
                    wallet.save()
            # --- End account update logic ---
            
            return JsonResponse({'success': True, 'transaction_id': transaction.id})
        except Exception as e:
            print(f"Error updating transaction: {e}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def delete_transaction_api(request, transaction_id):
    """API endpoint for deleting a transaction"""
    transaction = get_object_or_404(Transaction, id=transaction_id, user=request.user)
    
    if request.method == 'POST':
        try:
            # --- Revert the effect on account balance/usage before deleting ---
            account = transaction.transaction_account
            if account:
                amount_decimal = Decimal(str(transaction.amount))
                
                # Debit Account
                if hasattr(account, 'debitaccount'):
                    debit = account.debitaccount
                    if transaction.type == 'expense':
                        debit.balance += amount_decimal  # Add back the expense amount
                    else:  # income
                        debit.balance -= amount_decimal  # Subtract the income amount
                    debit.save()
                # Credit Account
                elif hasattr(account, 'creditaccount'):
                    credit = account.creditaccount
                    if transaction.type == 'expense':
                        credit.current_usage -= amount_decimal  # Reduce the usage
                    else:  # income (rare)
                        credit.current_usage += amount_decimal  # Increase the usage
                        if credit.current_usage > credit.credit_limit:
                            credit.current_usage = credit.credit_limit
                    credit.save()
                # Wallet
                elif hasattr(account, 'wallet'):
                    wallet = account.wallet
                    if transaction.type == 'expense':
                        wallet.balance += amount_decimal  # Add back the expense amount
                    else:  # income
                        wallet.balance -= amount_decimal  # Subtract the income amount
                    wallet.save()
            # --- End account update logic ---
                
            transaction.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def api_accounts(request):
    transaction_type = request.GET.get('transaction_type', 'expense')
    accounts = Account.objects.filter(user=request.user)
    
    # For income transactions, exclude credit accounts
    if transaction_type == 'income':
        accounts = accounts.exclude(creditaccount__isnull=False)
    
    account_data = []
    for account in accounts:
        # Get the specific account type instance
        try:
            if hasattr(account, 'debitaccount'):
                account_instance = account.debitaccount
                balance = account_instance.balance
                account_data.append({
                    'id': account.id,
                    'name': account.name,
                    'type': 'debit',
                    'balance': float(balance),
                    'maintaining_balance': float(account_instance.maintaining_balance) if account_instance.maintaining_balance is not None else 0
                })
            elif hasattr(account, 'creditaccount'):
                account_instance = account.creditaccount
                available_balance = account_instance.credit_limit - account_instance.current_usage
                account_data.append({
                    'id': account.id,
                    'name': account.name,
                    'type': 'credit',
                    'balance': float(available_balance),
                    'credit_limit': float(account_instance.credit_limit),
                    'current_usage': float(account_instance.current_usage)
                })
            elif hasattr(account, 'wallet'):
                account_instance = account.wallet
                balance = account_instance.balance
                account_data.append({
                    'id': account.id,
                    'name': account.name,
                    'type': 'wallet',
                    'balance': float(balance)
                })
            else:
                continue  # Skip if not a valid account type
        except Exception as e:
            print(f"Error processing account {account.id}: {e}")
            continue  # Skip if there's any error getting the account details
    
    return JsonResponse(account_data, safe=False)

@login_required
def api_account_balance(request, account_id):
    try:
        account = Account.objects.get(id=account_id, user=request.user)
        
        # Get the specific account type instance
        if hasattr(account, 'debitaccount'):
            account_instance = account.debitaccount
            balance = account_instance.balance
        elif hasattr(account, 'creditaccount'):
            account_instance = account.creditaccount
            balance = account_instance.credit_limit - account_instance.current_usage
        elif hasattr(account, 'wallet'):
            account_instance = account.wallet
            balance = account_instance.balance
        else:
            return JsonResponse({'error': 'Invalid account type'}, status=400)
            
        return JsonResponse({'balance': float(balance)})
    except Account.DoesNotExist:
        return JsonResponse({'error': 'Account not found'}, status=404)

@login_required
def charts_view(request):
    # Get the current date and calculate date ranges
    today = timezone.now().date()
    start_of_month = today.replace(day=1)
    start_of_year = today.replace(month=1, day=1)
    
    # Categories Analysis Data
    categories = Category.objects.filter(user=request.user)
    category_data = []
    
    for category in categories:
        total = Transaction.objects.filter(
            user=request.user,
            category=category,
            date__gte=start_of_month
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        if total != 0:  # Only include categories with transactions
            category_data.append({
                'category': category.name,
                'amount': float(total)
            })
    
    # Time Analysis Data (Last 6 months)
    time_data = {
        'labels': [],
        'income': [],
        'expenses': []
    }
    
    for i in range(6):
        month = today.replace(day=1) - timedelta(days=30*i)
        month_name = month.strftime('%b %Y')
        time_data['labels'].insert(0, month_name)
        
        # Get income for the month
        income = Transaction.objects.filter(
            user=request.user,
            type='income',
            date__year=month.year,
            date__month=month.month
        ).aggregate(total=Sum('amount'))['total'] or 0
        time_data['income'].insert(0, float(income))
        
        # Get expenses for the month
        expenses = Transaction.objects.filter(
            user=request.user,
            type='expense',
            date__year=month.year,
            date__month=month.month
        ).aggregate(total=Sum('amount'))['total'] or 0
        time_data['expenses'].insert(0, float(expenses))
    
    # Future Projections Data (Next 6 months)
    future_data = {
        'labels': [],
        'projected': []
    }
    
    # Calculate average monthly income and expenses from the last 3 months
    last_3_months_avg = Transaction.objects.filter(
        user=request.user,
        date__gte=today - timedelta(days=90)
    ).aggregate(avg=Avg('amount'))['avg'] or 0
    
    # Project next 6 months
    for i in range(1, 7):
        month = today.replace(day=1) + timedelta(days=30*i)
        month_name = month.strftime('%b %Y')
        future_data['labels'].append(month_name)
        future_data['projected'].append(float(last_3_months_avg))
    
    context = {
        'category_data': json.dumps(category_data),
        'time_data': json.dumps(time_data),
        'future_data': json.dumps(future_data)
    }
    
    return render(request, 'finances/charts.html', context)

@login_required
def charts_data_time(request):
    """API endpoint for time-based chart data"""
    # Get query parameters
    range_param = request.GET.get('range', 'month')
    interval = request.GET.get('interval', 'day')
    
    today = timezone.now().date()
    data = {
        'labels': [],
        'income': [],
        'expenses': []
    }
    
    if range_param == 'week':
        # Last 7 days
        for i in range(7):
            date = today - timedelta(days=i)
            data['labels'].insert(0, date.strftime('%a %d'))
            
            # Get income for the day
            income = Transaction.objects.filter(
                user=request.user,
                type='income',
                date=date
            ).aggregate(total=Sum('amount'))['total'] or 0
            data['income'].insert(0, float(income))
            
            # Get expenses for the day
            expenses = Transaction.objects.filter(
                user=request.user,
                type='expense',
                date=date
            ).aggregate(total=Sum('amount'))['total'] or 0
            data['expenses'].insert(0, float(expenses))
            
    elif range_param == 'month':
        if interval == 'day':
            # Last 30 days
            for i in range(30):
                date = today - timedelta(days=i)
                data['labels'].insert(0, date.strftime('%d'))
                
                # Get income for the day
                income = Transaction.objects.filter(
                    user=request.user,
                    type='income',
                    date=date
                ).aggregate(total=Sum('amount'))['total'] or 0
                data['income'].insert(0, float(income))
                
                # Get expenses for the day
                expenses = Transaction.objects.filter(
                    user=request.user,
                    type='expense',
                    date=date
                ).aggregate(total=Sum('amount'))['total'] or 0
                data['expenses'].insert(0, float(expenses))
                
        elif interval == 'week':
            # Last 4 weeks
            for i in range(4):
                end_date = today - timedelta(weeks=i)
                start_date = end_date - timedelta(days=6)
                data['labels'].insert(0, f'Week {4-i}')
                
                # Get income for the week
                income = Transaction.objects.filter(
                    user=request.user,
                    type='income',
                    date__gte=start_date,
                    date__lte=end_date
                ).aggregate(total=Sum('amount'))['total'] or 0
                data['income'].insert(0, float(income))
                
                # Get expenses for the week
                expenses = Transaction.objects.filter(
                    user=request.user,
                    type='expense',
                    date__gte=start_date,
                    date__lte=end_date
                ).aggregate(total=Sum('amount'))['total'] or 0
                data['expenses'].insert(0, float(expenses))
                
    elif range_param == 'year':
        # Last 12 months
        for i in range(12):
            month = today.replace(day=1) - timedelta(days=30*i)
            data['labels'].insert(0, month.strftime('%b'))
            
            # Get income for the month
            income = Transaction.objects.filter(
                user=request.user,
                type='income',
                date__year=month.year,
                date__month=month.month
            ).aggregate(total=Sum('amount'))['total'] or 0
            data['income'].insert(0, float(income))
            
            # Get expenses for the month
            expenses = Transaction.objects.filter(
                user=request.user,
                type='expense',
                date__year=month.year,
                date__month=month.month
            ).aggregate(total=Sum('amount'))['total'] or 0
            data['expenses'].insert(0, float(expenses))
    
    return JsonResponse(data)

@login_required
def charts_data_future(request):
    # Get the current date and calculate date ranges
    today = timezone.now()
    start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    # Calculate end date (6 months from now)
    if start_date.month + 6 > 12:
        end_date = start_date.replace(year=start_date.year + 1, month=start_date.month + 6 - 12)
    else:
        end_date = start_date.replace(month=start_date.month + 6)
    end_date = end_date.replace(day=1) - timedelta(days=1)
    end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Get all transactions (actual and scheduled) for the period
    transactions = get_transactions_with_scheduled(request.user, start_date, end_date)
    # Sort transactions by date
    transactions.sort(key=lambda t: t['date'])

    # Prepare labels (dates as strings)
    dates = [start_date.strftime('%Y-%m-%d')]
    for t in transactions:
        dates.append(t['date'].strftime('%Y-%m-%d'))

    # Calculate projected balance trend
    balances = []
    current_balance = Decimal('0')
    for account in Account.objects.filter(user=request.user):
        if hasattr(account, 'debitaccount'):
            current_balance += account.debitaccount.balance
        elif hasattr(account, 'wallet'):
            current_balance += account.wallet.balance
        elif hasattr(account, 'creditaccount'):
            current_balance -= account.creditaccount.current_usage
    balances.append(float(current_balance))
    for t in transactions:
        if t['type'] == 'income':
            current_balance += Decimal(str(t['amount']))
        else:
            current_balance -= Decimal(str(t['amount']))
        balances.append(float(current_balance))

    # For now, fill other arrays with zeros (same length as labels)
    n = len(dates)
    zeros = [0] * n
    data = {
        'labels': dates,
        'future_transactions': zeros,
        'scheduled_transactions': zeros,
        'debts_credits': zeros,
        'credit_card_payments': zeros,
        'projected': balances,
    }
    return JsonResponse(data)

def calculate_account_summaries(user):
    """Calculate account summaries for a user."""
    debit_accounts = DebitAccount.objects.filter(user=user)
    credit_accounts = CreditAccount.objects.filter(user=user)
    wallet_accounts = Wallet.objects.filter(user=user)
    
    debit_balance = Decimal('0.00')
    for account in debit_accounts:
        debit_balance += account.balance - account.maintaining_balance
    
    credit_balance = Decimal('0.00')
    for account in credit_accounts:
        credit_balance += account.credit_limit - account.current_usage
    
    wallet_balance = Decimal('0.00')
    for account in wallet_accounts:
        wallet_balance += account.balance
    
    return {
        'debit': {
            'accounts': list(debit_accounts),
            'balance': debit_balance
        },
        'credit': {
            'accounts': list(credit_accounts),
            'balance': credit_balance
        },
        'wallet': {
            'accounts': list(wallet_accounts),
            'balance': wallet_balance
        },
    }

def get_month_summary(user, month_date):
    """Get the income and expenses summary for a given month."""
    # Calculate the first and last day of the month
    first_day = month_date
    last_day = (month_date.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    
    # Get month's transactions
    month_income = Transaction.objects.filter(
        user=user,
        type='income',
        date__gte=first_day,
        date__lte=last_day
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    
    month_expenses = Transaction.objects.filter(
        user=user,
        type='expense',
        date__gte=first_day,
        date__lte=last_day
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    
    return month_income, month_expenses

def get_budget_warnings(user, start_date, end_date):
    """Get budget warnings for a given date range."""
    budgets = Budget.objects.filter(user=user, start_date__lte=end_date, end_date__gte=start_date)
    budget_warnings = []
    
    for budget in budgets:
        spent = Transaction.objects.filter(
            user=user,
            type='expense',
            category=budget.category,
            date__gte=start_date,
            date__lte=end_date
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        
        if spent > float(budget.amount) * 0.8:  # Warning at 80% of budget
            budget_warnings.append({
                'category': budget.category.name,
                'budget': budget.amount,
                'spent': spent,
                'percentage': round((spent / budget.amount) * 100)
            })
    
    return budget_warnings

@login_required
def batch_delete_transactions_api(request):
    """API endpoint for batch deleting multiple transactions"""
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Parse the request body
        data = json.loads(request.body)
        transaction_ids = data.get('transaction_ids', [])
        
        if not transaction_ids:
            return JsonResponse({'error': 'No transaction IDs provided'}, status=400)
        
        # Get transactions that belong to this user
        transactions = Transaction.objects.filter(
            user=request.user,
            id__in=transaction_ids
        )
        
        if not transactions:
            return JsonResponse({'error': 'No valid transactions found to delete'}, status=404)
        
        # Count how many were found
        found_count = transactions.count()
        
        # Delete the transactions
        transactions.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully deleted {found_count} transaction(s)',
            'deleted_count': found_count
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON in request body'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def create_test_scheduled_transaction(user):
    """
    Create a test scheduled transaction for testing purposes.
    Returns the created transaction.
    """
    # Get or create a test category
    category, _ = Category.objects.get_or_create(
        user=user,
        name='Test Category',
        type='expense',
        icon='bi-tag-fill',
        order=0
    )
    
    # Get or create a test account
    account, _ = DebitAccount.objects.get_or_create(
        user=user,
        name='Test Account',
        balance=Decimal('1000.00'),
        maintaining_balance=Decimal('10.00')
    )
    
    # Create a scheduled transaction that's due in 1 minutes
    scheduled = ScheduledTransaction.objects.create(
        user=user,
        name='Test Scheduled Transaction',
        category=category,
        transaction_type='expense',
        account=account,
        amount=Decimal('100.00'),
        date_scheduled=timezone.now() + timedelta(minutes=1),  # 1 minutes in the future
        repeat_type='monthly',
        repeats=3,
        note='Test transaction',
        status='scheduled'
    )
    
    return scheduled

@login_required
def test_scheduled_transactions(request):
    """
    Test endpoint to verify scheduled transaction processing.
    Creates a test transaction and processes it.
    """
    # Create a test scheduled transaction
    scheduled = create_test_scheduled_transaction(request.user)
    
    # Check if the scheduled time has passed
    if timezone.now() >= scheduled.date_scheduled:
        # Process scheduled transactions
        process_scheduled_transactions(request.user)
        
        # Get the updated scheduled transaction
        scheduled.refresh_from_db()
        
        # Get the created transaction if any
        created_transaction = Transaction.objects.filter(
            notes__contains=f"Scheduled transaction: {scheduled.name}"
        ).first()
        
        # Get the next occurrence if any
        next_occurrence = ScheduledTransaction.objects.filter(
            parent_transaction=scheduled,
            status='scheduled'
        ).first()
    else:
        # If scheduled time hasn't passed yet
        created_transaction = None
        next_occurrence = None
        messages.info(request, f"Test transaction is scheduled for {scheduled.date_scheduled}. Please wait until then to process it.")
    
    context = {
        'scheduled': scheduled,
        'created_transaction': created_transaction,
        'next_occurrence': next_occurrence,
    }
    
    return render(request, 'finances/test_scheduled.html', context)

@login_required
def resolve_scheduled_transaction(request, pk):
    """View to resolve a failed scheduled transaction."""
    scheduled = get_object_or_404(ScheduledTransaction, pk=pk, user=request.user)
    
    if scheduled.status != 'failed':
        return JsonResponse({
            'success': False,
            'error': 'Only failed transactions can be resolved.'
        })
    
    try:
        # First validate if the transaction can be processed with the current account state
        account = scheduled.account
        amount_decimal = Decimal(str(scheduled.amount))
        
        # Validate account balance/limits based on account type
        if hasattr(account, 'debitaccount'):
            debit = account.debitaccount
            if scheduled.transaction_type == 'expense':
                # Check if balance would go below maintaining balance
                if debit.balance - amount_decimal < (debit.maintaining_balance or 0):
                    scheduled.status = 'failed'
                    scheduled.save(update_fields=['status'])
                    return JsonResponse({
                        'success': False,
                        'error': 'Insufficient balance in debit account.'
                    })
        elif hasattr(account, 'creditaccount'):
            credit = account.creditaccount
            if scheduled.transaction_type == 'expense':
                # Check if credit limit would be exceeded
                if credit.current_usage + amount_decimal > credit.credit_limit:
                    scheduled.status = 'failed'
                    scheduled.save(update_fields=['status'])
                    return JsonResponse({
                        'success': False,
                        'error': 'Credit limit would be exceeded.'
                    })
        elif hasattr(account, 'wallet'):
            wallet = account.wallet
            if scheduled.transaction_type == 'expense' and wallet.balance < amount_decimal:
                scheduled.status = 'failed'
                scheduled.save(update_fields=['status'])
                return JsonResponse({
                    'success': False,
                    'error': 'Insufficient balance in wallet.'
                })
        
        # If validation passes, create the transaction
        transaction = Transaction.objects.create(
            user=request.user,
            title=scheduled.name,
            amount=scheduled.amount,
            date=scheduled.date_scheduled.date(),
            time=scheduled.date_scheduled.time(),
            type=scheduled.transaction_type,
            category=scheduled.category,
            subcategory=scheduled.subcategory,
            transaction_account=scheduled.account,
            notes=f"Scheduled transaction: {scheduled.name} (ID: {scheduled.id})"
        )
        
        # Update account balance/usage
        if hasattr(account, 'debitaccount'):
            debit = account.debitaccount
            if scheduled.transaction_type == 'expense':
                debit.balance -= amount_decimal
            else:  # income
                debit.balance += amount_decimal
            debit.save()
        elif hasattr(account, 'creditaccount'):
            credit = account.creditaccount
            if scheduled.transaction_type == 'expense':
                credit.current_usage += amount_decimal
            else:  # income (rare)
                credit.current_usage -= amount_decimal
                if credit.current_usage < 0:
                    credit.current_usage = 0
            credit.save()
        elif hasattr(account, 'wallet'):
            wallet = account.wallet
            if scheduled.transaction_type == 'expense':
                wallet.balance -= amount_decimal
            else:  # income
                wallet.balance += amount_decimal
            wallet.save()
        
        # Update the scheduled transaction status to completed
        scheduled.status = 'completed'
        scheduled.save(update_fields=['status'])
        
        # If this is a recurring transaction, create the next occurrence
        if scheduled.repeats != 1:
            # Only create next if repeats is infinite or > 1
            if scheduled.repeats == 0:
                next_repeats = 0
            elif scheduled.repeats > 1:
                next_repeats = scheduled.repeats - 1
            else:
                next_repeats = None
            if next_repeats is not None:
                next_date = scheduled.date_scheduled
                if scheduled.repeat_type == 'daily':
                    next_date += timedelta(days=1)
                elif scheduled.repeat_type == 'weekly':
                    next_date += timedelta(weeks=1)
                elif scheduled.repeat_type == 'monthly':
                    if next_date.month == 12:
                        next_date = next_date.replace(year=next_date.year + 1, month=1)
                    else:
                        next_date = next_date.replace(month=next_date.month + 1)
                elif scheduled.repeat_type == 'yearly':
                    next_date = next_date.replace(year=next_date.year + 1)
                ScheduledTransaction.objects.create(
                    user=request.user,
                    name=scheduled.name,
                    category=scheduled.category,
                    subcategory=scheduled.subcategory,
                    transaction_type=scheduled.transaction_type,
                    account=scheduled.account,
                    amount=scheduled.amount,
                    date_scheduled=next_date,
                    repeat_type=scheduled.repeat_type,
                    repeats=next_repeats,
                    note=scheduled.note,
                    status='scheduled'
                )
        return JsonResponse({
            'success': True,
            'message': 'Transaction resolved successfully.'
        })
        
    except Exception as e:
        # If any error occurs, mark the transaction as failed
        scheduled.status = 'failed'
        scheduled.save(update_fields=['status'])
        return JsonResponse({
            'error': str(e),
            'labels': [],
            'projected': [],
            'future_transactions': [],
            'scheduled_transactions': [],
            'debts_credits': [],
            'credit_card_payments': []
        }, status=500)

@login_required
def export_data(request):
    if request.method == 'POST':
        form = ExportForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            format = form.cleaned_data['format']
            include_fields = form.cleaned_data['include_fields']
            
            # Get transactions with related fields
            transactions = Transaction.objects.filter(
                user=request.user,
                date__gte=start_date,
                date__lte=end_date
            ).select_related('category', 'subcategory', 'transaction_account').order_by('-date')
            
            if format == 'csv':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="transactions_{start_date}_{end_date}.csv"'
                
                writer = csv.writer(response)
                writer.writerow(include_fields)  # Write header row
                
                for transaction in transactions:
                    row = []
                    for field in include_fields:
                        if field == 'id':
                            row.append(transaction.id)
                        elif field == 'title':
                            row.append(transaction.title)
                        elif field == 'amount':
                            row.append(transaction.amount)
                        elif field == 'currency':
                            row.append(transaction.currency)
                        elif field == 'date':
                            row.append(transaction.date)
                        elif field == 'time':
                            row.append(transaction.time)
                        elif field == 'type':
                            row.append(transaction.type)
                        elif field == 'category':
                            row.append(transaction.category.name if transaction.category else '')
                        elif field == 'subcategory':
                            row.append(transaction.subcategory.name if transaction.subcategory else '')
                        elif field == 'account':
                            row.append(transaction.transaction_account.name if transaction.transaction_account else '')
                        elif field == 'notes':
                            row.append(transaction.notes or '')
                        elif field == 'photo':
                            row.append(request.build_absolute_uri(transaction.photo.url) if transaction.photo else '')
                    writer.writerow(row)
                
                return response
            
            elif format == 'json':
                data = []
                for transaction in transactions:
                    transaction_data = {}
                    for field in include_fields:
                        if field == 'id':
                            transaction_data['id'] = transaction.id
                        elif field == 'title':
                            transaction_data['title'] = transaction.title
                        elif field == 'amount':
                            transaction_data['amount'] = float(transaction.amount)
                        elif field == 'currency':
                            transaction_data['currency'] = transaction.currency
                        elif field == 'date':
                            transaction_data['date'] = transaction.date.isoformat()
                        elif field == 'time':
                            transaction_data['time'] = transaction.time.strftime('%H:%M:%S')
                        elif field == 'type':
                            transaction_data['type'] = transaction.type
                        elif field == 'category':
                            transaction_data['category'] = transaction.category.name if transaction.category else None
                        elif field == 'subcategory':
                            transaction_data['subcategory'] = transaction.subcategory.name if transaction.subcategory else None
                        elif field == 'account':
                            transaction_data['account'] = transaction.transaction_account.name if transaction.transaction_account else None
                        elif field == 'notes':
                            transaction_data['notes'] = transaction.notes
                        elif field == 'photo':
                            transaction_data['photo'] = request.build_absolute_uri(transaction.photo.url) if transaction.photo else None
                    data.append(transaction_data)
                
                response = HttpResponse(json.dumps(data, indent=2), content_type='application/json')
                response['Content-Disposition'] = f'attachment; filename="transactions_{start_date}_{end_date}.json"'
                return response
            
            elif format == 'xlsx':
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Transactions'
                
                # Write header row with styling
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                
                for col, field in enumerate(include_fields, 1):
                    cell = ws.cell(row=1, column=col, value=field)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Write data rows
                for row, transaction in enumerate(transactions, 2):
                    for col, field in enumerate(include_fields, 1):
                        if field == 'id':
                            value = transaction.id
                        elif field == 'title':
                            value = transaction.title
                        elif field == 'amount':
                            value = float(transaction.amount)
                        elif field == 'currency':
                            value = transaction.currency
                        elif field == 'date':
                            value = transaction.date
                        elif field == 'time':
                            value = transaction.time
                        elif field == 'type':
                            value = transaction.type
                        elif field == 'category':
                            value = transaction.category.name if transaction.category else ''
                        elif field == 'subcategory':
                            value = transaction.subcategory.name if transaction.subcategory else ''
                        elif field == 'account':
                            value = transaction.transaction_account.name if transaction.transaction_account else ''
                        elif field == 'notes':
                            value = transaction.notes or ''
                        elif field == 'photo':
                            value = request.build_absolute_uri(transaction.photo.url) if transaction.photo else ''
                        ws.cell(row=row, column=col, value=value)
                
                # Auto-adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width
                
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="transactions_{start_date}_{end_date}.xlsx"'
                wb.save(response)
                return response
    else:
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
        form = ExportForm(initial={'start_date': start_date, 'end_date': end_date})
    
    return render(request, 'finances/export_data.html', {'form': form})

@login_required
def import_data(request):
    if request.method == 'POST':
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            format = form.cleaned_data['format']
            duplicate_handling = form.cleaned_data['duplicate_handling']
            date_format = form.cleaned_data['date_format'] or '%Y-%m-%d'
            
            try:
                if format == 'csv':
                    decoded_file = file.read().decode('utf-8')
                    csv_data = csv.reader(decoded_file.splitlines(), delimiter=',')
                    headers = next(csv_data)  # Skip header row
                    
                    # Process each row
                    for row in csv_data:
                        if len(row) >= 5:  # Ensure minimum required fields
                            title = row[0]
                            amount = float(row[1])
                            date = datetime.strptime(row[2], date_format).date()
                            transaction_type = row[3]
                            category_name = row[4]
                            notes = row[5] if len(row) > 5 else ''
                            
                            # Get or create category
                            category = None
                            if category_name:
                                category, _ = Category.objects.get_or_create(
                                    user=request.user,
                                    name=category_name,
                                    type=transaction_type
                                )
                            
                            # Handle duplicates based on user preference
                            if duplicate_handling == 'skip':
                                if Transaction.objects.filter(
                                    user=request.user,
                                    title=title,
                                    amount=amount,
                                    date=date,
                                    type=transaction_type
                                ).exists():
                                    continue
                            
                            # Create transaction
                            Transaction.objects.create(
                                user=request.user,
                                title=title,
                                amount=amount,
                                date=date,
                                type=transaction_type,
                                category=category,
                                notes=notes
                            )
                    
                elif format == 'json':
                    data = json.loads(file.read().decode('utf-8'))
                    for item in data:
                        # Handle duplicates based on user preference
                        if duplicate_handling == 'skip':
                            if Transaction.objects.filter(
                                user=request.user,
                                title=item['title'],
                                amount=item['amount'],
                                date=datetime.strptime(item['date'], date_format).date(),
                                type=item['type']
                            ).exists():
                                continue
                        
                        # Get or create category
                        category = None
                        if 'category' in item and item['category']:
                            category, _ = Category.objects.get_or_create(
                                user=request.user,
                                name=item['category'],
                                type=item['type']
                            )
                        
                        # Create transaction
                        Transaction.objects.create(
                            user=request.user,
                            title=item['title'],
                            amount=item['amount'],
                            date=datetime.strptime(item['date'], date_format).date(),
                            time=datetime.strptime(item['time'], '%H:%M:%S').time() if 'time' in item else None,
                            type=item['type'],
                            category=category,
                            notes=item.get('notes', '')
                        )
                
                elif format == 'xlsx':
                    import openpyxl
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    
                    for row in ws.iter_rows(min_row=2):
                        row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
                        
                        # Handle duplicates based on user preference
                        if duplicate_handling == 'skip':
                            if Transaction.objects.filter(
                                user=request.user,
                                title=row_data['title'],
                                amount=float(row_data['amount']),
                                date=datetime.strptime(row_data['date'], date_format).date(),
                                type=row_data['type']
                            ).exists():
                                continue
                        
                        # Get or create category
                        category = None
                        if 'category' in row_data and row_data['category']:
                            category, _ = Category.objects.get_or_create(
                                user=request.user,
                                name=row_data['category'],
                                type=row_data['type']
                            )
                        
                        # Create transaction
                        Transaction.objects.create(
                            user=request.user,
                            title=row_data['title'],
                            amount=float(row_data['amount']),
                            date=datetime.strptime(row_data['date'], date_format).date(),
                            time=datetime.strptime(row_data['time'], '%H:%M:%S').time() if 'time' in row_data else None,
                            type=row_data['type'],
                            category=category,
                            notes=row_data.get('notes', '')
                        )
                
                messages.success(request, 'Transactions imported successfully!')
                return redirect('transactions')
            
            except Exception as e:
                messages.error(request, f'Error importing file: {str(e)}')
    else:
        form = ImportForm()
    
    return render(request, 'finances/import_data.html', {'form': form})

def import_export_data(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    export_form = ExportForm(user=request.user)
    import_form = ImportForm()
    
    if request.method == 'POST':
        operation = request.POST.get('operation')
        
        if operation == 'export':
            export_form = ExportForm(request.POST, user=request.user)
            if export_form.is_valid():
                start_date = export_form.cleaned_data['start_date']
                end_date = export_form.cleaned_data['end_date']
                format = export_form.cleaned_data['format']
                separator = export_form.cleaned_data['separator']
                account = export_form.cleaned_data['account']
                include_income = export_form.cleaned_data['include_income']
                include_expenses = export_form.cleaned_data['include_expenses']
                
                # Build transaction filter
                transaction_filter = {
                    'user': request.user,
                    'date__range': [start_date, end_date]
                }
                
                # Filter by account if specific account selected
                if account != 'all':
                    transaction_filter['transaction_account_id'] = account
                
                # Filter by transaction type
                transaction_types = []
                if include_income:
                    transaction_types.append('income')
                if include_expenses:
                    transaction_types.append('expense')
                if transaction_types:
                    transaction_filter['type__in'] = transaction_types
                
                transactions = Transaction.objects.filter(**transaction_filter).order_by('date')
                
                # Prepare headers and data
                headers = ['Date', 'Title', 'Amount', 'Type', 'Category', 'Account', 'Notes']
                
                if format == 'csv':
                    response = HttpResponse(content_type='text/csv')
                    response['Content-Disposition'] = f'attachment; filename="transactions_{start_date}_{end_date}.csv"'
                    
                    writer = csv.writer(response, delimiter=separator)
                    writer.writerow(headers)
                    
                    for transaction in transactions:
                        writer.writerow([
                            transaction.date,
                            transaction.title,
                            transaction.amount,
                            transaction.type,
                            transaction.category.name if transaction.category else '',
                            transaction.transaction_account.name if transaction.transaction_account else '',
                            transaction.notes or ''
                        ])
                    
                    return response
                    
                elif format == 'xlsx':
                    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename="transactions_{start_date}_{end_date}.xlsx"'
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = 'Transactions'
                    
                    # Write headers
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                    
                    # Write data
                    for row, transaction in enumerate(transactions, 2):
                        ws.cell(row=row, column=1, value=transaction.date)
                        ws.cell(row=row, column=2, value=transaction.title)
                        ws.cell(row=row, column=3, value=float(transaction.amount))
                        ws.cell(row=row, column=4, value=transaction.type)
                        ws.cell(row=row, column=5, value=transaction.category.name if transaction.category else '')
                        ws.cell(row=row, column=6, value=transaction.transaction_account.name if transaction.transaction_account else '')
                        ws.cell(row=row, column=7, value=transaction.notes or '')
                    
                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column[0].column_letter].width = adjusted_width
                    
                    wb.save(response)
                    return response
                    
        elif operation == 'import':
            import_form = ImportForm(request.POST, request.FILES)
            if import_form.is_valid():
                file = import_form.cleaned_data['file']
                duplicate_handling = import_form.cleaned_data['duplicate_handling']
                
                try:
                    # Determine file format from extension
                    file_format = file.name.split('.')[-1].lower()
                    
                    if file_format == 'csv':
                        decoded_file = file.read().decode('utf-8').splitlines()
                        reader = csv.DictReader(decoded_file)
                        transactions_data = list(reader)
                    elif file_format == 'xlsx':
                        wb = openpyxl.load_workbook(file)
                        ws = wb.active
                        headers = [cell.value for cell in ws[1]]
                        transactions_data = []
                        for row in ws.iter_rows(min_row=2):
                            transaction = {}
                            for header, cell in zip(headers, row):
                                transaction[header] = cell.value
                            transactions_data.append(transaction)
                    
                    success_count = 0
                    error_count = 0
                    errors = []
                    
                    expected_headers = ['Date', 'Title', 'Amount', 'Type', 'Category', 'Account', 'Notes']
                    if not all(header in (transactions_data[0].keys() if transactions_data else []) for header in expected_headers):
                        raise ValueError('Invalid file format. This file does not match the export format from this system.')
                    
                    for data in transactions_data:
                        try:
                            # Parse date (our export format uses YYYY-MM-DD)
                            date_str = str(data['Date'])  # Convert to string first
                            # Handle both date and datetime formats
                            try:
                                date = datetime.strptime(date_str, '%Y-%m-%d').date()
                            except ValueError:
                                # If the string includes time, try parsing with time
                                date = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S').date()
                            
                            # Get or create category if provided
                            category = None
                            if data['Category']:
                                category, _ = Category.objects.get_or_create(
                                    user=request.user,
                                    name=data['Category'],
                                    defaults={'type': data['Type']}
                                )
                            
                            # Get or create account if provided
                            account = None
                            if data['Account']:
                                account, _ = Account.objects.get_or_create(
                                    user=request.user,
                                    name=data['Account']
                                )
                            
                            # Check for duplicates
                            existing_transaction = Transaction.objects.filter(
                                user=request.user,
                                title=data['Title'],
                                amount=Decimal(str(data['Amount'])),
                                date=date
                            ).first()
                            
                            if existing_transaction:
                                if duplicate_handling == 'skip':
                                    continue
                                elif duplicate_handling == 'update':
                                    existing_transaction.type = data['Type']
                                    existing_transaction.category = category
                                    existing_transaction.transaction_account = account
                                    existing_transaction.notes = data['Notes']
                                    existing_transaction.save()
                                    success_count += 1
                                    continue
                            
                            # Create new transaction
                            Transaction.objects.create(
                                user=request.user,
                                title=data['Title'],
                                amount=Decimal(str(data['Amount'])),
                                date=date,
                                type=data['Type'],
                                category=category,
                                transaction_account=account,
                                notes=data['Notes']
                            )
                            success_count += 1
                            
                        except Exception as e:
                            error_count += 1
                            errors.append(str(e))
                    
                    if error_count > 0:
                        messages.warning(request, f'Import completed with {success_count} successful imports and {error_count} errors.')
                        if errors:
                            messages.error(request, 'Errors: ' + ', '.join(errors))
                    else:
                        messages.success(request, f'Successfully imported {success_count} transactions.')
                    
                except ValueError as ve:
                    messages.error(request, str(ve))
                except Exception as e:
                    messages.error(request, f'Error processing file: {str(e)}')
    
    return render(request, 'finances/import_export.html', {
        'export_form': export_form,
        'import_form': import_form
    })
